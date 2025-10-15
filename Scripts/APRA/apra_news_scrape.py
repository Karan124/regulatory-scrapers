import os
import json
import time
import hashlib
import logging
import requests
import urllib3
import random
from datetime import datetime, timezone
from typing import List, Dict, Optional, Set
from urllib.parse import urljoin, urlparse
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import PyPDF2
import io
import re

try:
    from selenium_stealth import stealth
    STEALTH_AVAILABLE = True
except ImportError:
    STEALTH_AVAILABLE = False

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel extraction will be skipped")

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configuration
DATA_DIR = "data"
JSON_PATH = os.path.join(DATA_DIR, "apra_news.json")
BASE_URL = "https://www.apra.gov.au"
START_URL = "https://www.apra.gov.au/news-and-publications"

# Rotating User-Agents for stealth
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0"
]

# Settings
MAX_PAGES = 1
PAGE_LOAD_TIMEOUT = 45
ARTICLE_TIMEOUT = 30
MIN_CONTENT_LENGTH = 50
MIN_HEADLINE_LENGTH = 10

# Link validation rules
EXCLUDED_SOCIAL_DOMAINS = ['twitter.com', 'x.com', 'linkedin.com', 'youtube.com', 'facebook.com']
EXCLUDED_FILE_EXTENSIONS = ['.jpg', '.jpeg', '.png', '.gif', '.css', '.js', '.ico']
ATTACHMENT_EXTENSIONS = ['.pdf', '.xlsx', '.xls']  # Only PDF and Excel files
EXCLUDED_URLS = [
    "https://www.apra.gov.au/newsletter-signup",
    "https://www.apra.gov.au/news-and-publications",
    "https://www.apra.gov.au/",
    "https://www.apra.gov.au"
]
EXCLUDED_URL_PATTERNS = [
    "news-and-publications?industry",
    "news-and-publications/page",
    "/news-and-publications?",
    "/news-and-publications/"
]

def setup_logging():
    os.makedirs(DATA_DIR, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s %(levelname)s: %(message)s',
        handlers=[
            logging.FileHandler(os.path.join(DATA_DIR, 'scraper.log')),
            logging.StreamHandler()
        ]
    )

def setup_stealth_driver():
    """Setup Chrome driver with enhanced stealth"""
    chrome_options = Options()
    
    # Enhanced stealth options
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    # Random window size and user agent
    width = random.randint(1200, 1920)
    height = random.randint(800, 1080)
    chrome_options.add_argument(f"--window-size={width},{height}")
    chrome_options.add_argument(f'--user-agent={random.choice(USER_AGENTS)}')
    
    # Performance optimizations
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-plugins")
    chrome_options.add_argument("--disable-background-networking")
    chrome_options.add_argument("--disable-default-apps")
    chrome_options.add_argument("--no-first-run")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    if STEALTH_AVAILABLE:
        stealth(driver, languages=["en-US", "en"], vendor="Google Inc.", platform="Win32")
    
    # Remove webdriver indicators
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.execute_script("Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]})")
    driver.execute_script("Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']})")
    
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    return driver

def create_session():
    session = requests.Session()
    session.headers.update({
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive"
    })
    session.verify = False
    return session

def human_delay():
    """Human-like random delays"""
    time.sleep(random.uniform(2, 5))

def generate_hash_id(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]

def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def count_tokens(text: str) -> int:
    """Rough token count estimation (4 characters = 1 token approximately)"""
    return len(text) // 4

def is_valid_link(link_url: str, original_url: str) -> bool:
    """Validate if a link should be processed as one-level-down"""
    try:
        if not link_url.startswith(('http://', 'https://')):
            return False
        
        # Check excluded URLs
        if link_url in EXCLUDED_URLS:
            return False
            
        # Check excluded URL patterns
        if any(pattern in link_url for pattern in EXCLUDED_URL_PATTERNS):
            return False
        
        if link_url == original_url:
            return False
            
        clean_link = link_url.split('#')[0].split('?')[0]
        clean_original = original_url.split('#')[0].split('?')[0]
        if clean_link == clean_original:
            return False
        
        parsed_url = urlparse(link_url)
        domain = parsed_url.netloc.lower()
        if any(social in domain for social in EXCLUDED_SOCIAL_DOMAINS):
            return False
        
        path = parsed_url.path.lower()
        if any(ext in path for ext in EXCLUDED_FILE_EXTENSIONS):
            return False
            
        if link_url.lower().startswith(('mailto:', 'javascript:')):
            return False
        
        return True
    except:
        return False

def extract_hyperlinks(soup: BeautifulSoup, base_url: str) -> List[str]:
    """Extract all hyperlinks from content areas"""
    links = []
    
    # Focus on content areas, avoid navigation
    content_areas = soup.select('.rich-text, main, article, .content, .page-content')
    if not content_areas:
        content_areas = [soup]
    
    for area in content_areas:
        for link in area.find_all('a', href=True):
            href = link.get('href')
            if href:
                full_url = urljoin(base_url, href)
                links.append(full_url)
    
    return list(set(links))

def get_attachment_info(url: str, session: requests.Session) -> Optional[Dict]:
    """Get attachment metadata"""
    try:
        response = session.head(url, timeout=15)
        response.raise_for_status()
        
        filename = os.path.basename(urlparse(url).path) or 'unknown_file'
        size = response.headers.get('content-length', 'unknown')
        
        return {
            'url': url,
            'filename': filename,
            'content_type': response.headers.get('content-type', 'unknown'),
            'size': int(size) if size.isdigit() else size
        }
    except:
        return None

def find_attachments(soup: BeautifulSoup, base_url: str, session: requests.Session) -> List[Dict]:
    """Find all PDF and Excel attachments from a page"""
    attachments = []
    
    for link in soup.find_all('a', href=True):
        href = link.get('href')
        if href:
            full_url = urljoin(base_url, href)
            path = urlparse(full_url).path.lower()
            
            # Only include PDF and Excel files
            if any(ext in path for ext in ATTACHMENT_EXTENSIONS):
                attachment_info = get_attachment_info(full_url, session)
                if attachment_info:
                    attachments.append(attachment_info)
    
    return attachments

def extract_excel_text(excel_url: str, session: requests.Session) -> str:
    """Extract text from Excel files"""
    if not EXCEL_AVAILABLE:
        logging.warning(f"openpyxl not available - skipping Excel extraction for {excel_url}")
        return ""
        
    try:
        response = session.get(excel_url, timeout=30)
        response.raise_for_status()
        
        from io import BytesIO
        wb = load_workbook(BytesIO(response.content), data_only=True)
        text_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = f"Sheet: {sheet_name}\n"
            
            # Extract all cell values
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    sheet_text += " | ".join(row_text) + "\n"
            
            text_parts.append(sheet_text)
        
        return '\n\n'.join(text_parts)
    except Exception as e:
        logging.warning(f"Error extracting Excel text from {excel_url}: {e}")
        return ""

def extract_pdf_text(pdf_url: str, session: requests.Session) -> str:
    """Extract text from PDF"""
    try:
        response = session.get(pdf_url, timeout=30)
        response.raise_for_status()
        
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(response.content))
        text_parts = []
        
        for page in pdf_reader.pages:
            try:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(clean_text(page_text))
            except:
                continue
        
        return '\n\n'.join(text_parts)
    except Exception as e:
        logging.warning(f"Error extracting PDF text from {pdf_url}: {e}")
        return ""

def remove_breadcrumbs(soup: BeautifulSoup) -> BeautifulSoup:
    """Remove breadcrumb navigation and other unwanted content from soup"""
    # Remove breadcrumb elements
    for breadcrumb in soup.select('.breadcrumb, .breadcrumbs, nav[aria-label="breadcrumb"], .nav-breadcrumb'):
        breadcrumb.decompose()
    
    # Remove specific breadcrumb text patterns
    breadcrumb_patterns = [
        r'Breadcrumb\s+Home\s+News and publications',
        r'Home\s+News and publications\s+News and Publications',
        r'Show AllCorporate publicationsInsightMedia ReleasesNewsOpening statementsSpeeches',
        r'Show AllCorporate publications',
        r'Breadcrumb',
        r'News and Publications\s+Show All'
    ]
    
    for pattern in breadcrumb_patterns:
        for elem in soup.find_all(text=re.compile(pattern, re.I)):
            if elem.parent:
                # Remove the parent element containing the breadcrumb text
                parent = elem.parent
                while parent and parent.name != 'body':
                    # Check if this element only contains breadcrumb-like content
                    text_content = clean_text(parent.get_text())
                    if any(re.search(bp, text_content, re.I) for bp in breadcrumb_patterns):
                        parent.decompose()
                        break
                    parent = parent.parent
                else:
                    # Just remove the text node itself
                    elem.extract()
    
    # Remove navigation and filter elements
    for nav_elem in soup.select('nav, .navigation, .filter, .filters, .pagination'):
        nav_elem.decompose()
    
    return soup

def extract_content_from_page(url: str, driver, session: requests.Session, page_type: str = "main") -> Dict:
    """Extract comprehensive content from any page"""
    try:
        logging.info(f"Extracting content from {page_type} page: {url}")
        
        # Load page with driver
        driver.get(url)
        WebDriverWait(driver, ARTICLE_TIMEOUT).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(3)
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        # Remove breadcrumbs if this is a linked page
        if page_type != "main":
            soup = remove_breadcrumbs(soup)
        
        # Extract content with comprehensive selectors
        content_parts = []
        
        content_selectors = [
            '.rich-text',           # APRA primary content area
            'main .content',        # Main content wrapper
            'article .content',     # Article content
            '.article-body',        # Article body
            '.news-content',        # News content
            '.page-content',        # Page content
            'main',                 # Main element
            'article'               # Article element
        ]
        
        found_substantial_content = False
        
        for selector in content_selectors:
            elements = soup.select(selector)
            if elements:
                for elem in elements:
                    # Extract all text elements including tables
                    for text_elem in elem.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'li', 'td', 'th', 'div', 'span']):
                        text = clean_text(text_elem.get_text())
                        if text and len(text) > 5:
                            # Format based on element type
                            if text_elem.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                                content_parts.append(f"\n## {text}\n")
                            elif text_elem.name == 'li':
                                content_parts.append(f"• {text}")
                            elif text_elem.name in ['td', 'th']:
                                content_parts.append(f"| {text}")  # Table formatting
                            else:
                                content_parts.append(text)
                
                if content_parts:
                    found_substantial_content = True
                    break
        
        # Fallback: extract all visible text if no structured content found
        if not found_substantial_content:
            # Remove unwanted elements
            for elem in soup(['script', 'style', 'nav', 'footer', 'header', 'aside', '.navigation', '.menu', '.breadcrumb']):
                elem.decompose()
            
            main_text = clean_text(soup.get_text())
            if main_text:
                content_parts = [main_text]
        
        main_content = '\n\n'.join(content_parts) if content_parts else ''
        
        # Find attachments (only PDF and Excel)
        attachments = find_attachments(soup, url, session)
        
        # Extract attachment text content and add to main content
        attachment_text_content = ""
        for attachment in attachments:
            filename = attachment['filename'].lower()
            if filename.endswith('.pdf'):
                pdf_text = extract_pdf_text(attachment['url'], session)
                if pdf_text:
                    attachment_text_content += f"\n\n--- PDF: {attachment['filename']} ---\n{pdf_text}"
            elif filename.endswith(('.xlsx', '.xls')):
                excel_text = extract_excel_text(attachment['url'], session)
                if excel_text:
                    attachment_text_content += f"\n\n--- Excel: {attachment['filename']} ---\n{excel_text}"
        
        # Combine all content
        all_content = main_content + attachment_text_content
        
        # Extract metadata if this is the main page
        headline = "Unknown"
        article_type = "Unknown"
        published_date = "Unknown"
        
        if page_type == "main":
            # Extract headline
            h1_elem = soup.select_one('h1, main h1, article h1')
            if h1_elem:
                headline = clean_text(h1_elem.get_text())
            
            # Extract article type
            type_selectors = [
                '.field-field-category',  # APRA specific
                '.tile__subject .field',  # APRA specific
                '.category',
                '.type',
                '[class*="category"]',
                '[class*="type"]'
            ]
            
            for selector in type_selectors:
                type_elem = soup.select_one(selector)
                if type_elem:
                    type_text = clean_text(type_elem.get_text())
                    if type_text and len(type_text) < 50:  # Reasonable length
                        article_type = type_text
                        break
            
            # Extract published date
            date_elem = soup.select_one('time[datetime]')
            if date_elem:
                published_date = clean_text(date_elem.get_text())
            else:
                # Try alternative date selectors
                for date_selector in ['.date', '.published', '[class*="date"]']:
                    date_elem = soup.select_one(date_selector)
                    if date_elem:
                        date_text = clean_text(date_elem.get_text())
                        if date_text and len(date_text) < 50:
                            published_date = date_text
                            break
        
        result = {
            'url': url,
            'content': all_content,
            'attachments': attachments,
            'token_count': count_tokens(all_content),
            'content_length': len(all_content)
        }
        
        if page_type == "main":
            result['headline'] = headline
            result['article_type'] = article_type
            result['published_date'] = published_date
        
        logging.info(f"Extracted {len(all_content)} characters, {len(attachments)} attachments from {page_type} page")
        
        return result
        
    except Exception as e:
        logging.error(f"Error extracting content from {url}: {e}")
        return {
            'url': url,
            'content': '',
            'attachments': [],
            'token_count': 0,
            'error': str(e)
        }

def process_article_with_links(url: str, driver, session: requests.Session) -> Dict:
    """Process main article and all valid one-level-down links"""
    try:
        logging.info(f"Processing article with one-level-down links: {url}")
        
        # Extract content from main article
        main_result = extract_content_from_page(url, driver, session, "main")
        
        if 'error' in main_result:
            return {
                'url': url,
                'headline': 'Unknown',
                'article_type': 'Unknown',
                'published_date': 'Unknown',
                'main_content': '',
                'linked_pages': [],
                'all_attachments': [],
                'total_token_count': 0,
                'error': main_result['error']
            }
        
        # Extract hyperlinks and process valid ones
        linked_results = []
        all_attachments = main_result['attachments'].copy()
        
        try:
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            hyperlinks = extract_hyperlinks(soup, url)
            valid_links = [link for link in hyperlinks if is_valid_link(link, url)]
            
            logging.info(f"Processing {len(valid_links)} one-level-down links")
            
            # Process up to 10 valid links
            for i, link in enumerate(valid_links[:10]):
                logging.info(f"Processing linked page {i+1}: {link}")
                
                link_result = extract_content_from_page(link, driver, session, f"linked-{i+1}")
                if 'error' not in link_result and link_result['content']:
                    linked_results.append({
                        'url': link,
                        'content': link_result['content'],
                        'attachments': link_result['attachments'],
                        'token_count': link_result['token_count']
                    })
                    
                    # Add attachments from linked page
                    all_attachments.extend(link_result['attachments'])
                
                human_delay()  # Stealth delay between requests
                
        except Exception as e:
            logging.warning(f"Error processing links: {e}")
        
        # Deduplicate attachments
        seen_attachments = set()
        unique_attachments = []
        for att in all_attachments:
            key = f"{att['url']}|{att['filename']}"
            if key not in seen_attachments:
                seen_attachments.add(key)
                unique_attachments.append(att)
        
        # Calculate total tokens from all content
        main_tokens = main_result['token_count']
        linked_tokens = sum(result['token_count'] for result in linked_results)
        total_tokens = main_tokens + linked_tokens
        
        return {
            'url': url,
            'headline': main_result.get('headline', 'Unknown'),
            'article_type': main_result.get('article_type', 'Unknown'),
            'published_date': main_result.get('published_date', 'Unknown'),
            'main_content': main_result['content'],
            'linked_pages': linked_results,
            'all_attachments': unique_attachments,
            'total_token_count': total_tokens,
            'main_token_count': main_tokens,
            'linked_token_count': linked_tokens,
            'linked_pages_count': len(linked_results),
            'attachments_count': len(unique_attachments)
        }
        
    except Exception as e:
        logging.error(f"Error in process_article_with_links: {e}")
        return {
            'url': url,
            'headline': 'Unknown',
            'article_type': 'Unknown', 
            'published_date': 'Unknown',
            'main_content': '',
            'linked_pages': [],
            'all_attachments': [],
            'total_token_count': 0,
            'error': str(e)
        }

def extract_article_metadata(link_element, base_url: str) -> Optional[Dict]:
    """Extract article metadata including type and published date"""
    try:
        href = link_element.get('href')
        if not href or any(param in href for param in ['?industry=', '?page=', '?tags=']):
            return None
            
        full_url = urljoin(base_url, href)
        
        # Find container with article info
        container = link_element.find_parent(['article', 'div', 'li'])
        if not container:
            return None
        
        # Extract headline
        headline = clean_text(link_element.get_text())
        if not headline or len(headline) < MIN_HEADLINE_LENGTH:
            for selector in ['h1', 'h2', 'h3', 'h4', '.title']:
                elem = container.select_one(selector)
                if elem:
                    headline = clean_text(elem.get_text())
                    if len(headline) >= MIN_HEADLINE_LENGTH:
                        break
        
        if not headline or len(headline) < MIN_HEADLINE_LENGTH:
            return None
        
        # Extract article type from multiple possible locations
        article_type = "Unknown"
        
        # Look for category/type in various selectors
        type_selectors = [
            '.field-field-category',  # APRA specific
            '.tile__subject .field',  # APRA specific
            '.category',
            '.type',
            '[class*="category"]',
            '[class*="type"]'
        ]
        
        for selector in type_selectors:
            type_elem = container.select_one(selector)
            if type_elem:
                type_text = clean_text(type_elem.get_text())
                if type_text and len(type_text) < 50:  # Reasonable length
                    article_type = type_text
                    break
        
        # Extract published date
        published_date = "Unknown"
        date_elem = container.select_one('time[datetime]')
        if date_elem:
            published_date = clean_text(date_elem.get_text())
        else:
            # Try alternative date patterns
            for date_selector in ['.date', '.published', '[class*="date"]']:
                date_elem = container.select_one(date_selector)
                if date_elem:
                    date_text = clean_text(date_elem.get_text())
                    if date_text and len(date_text) < 50:
                        published_date = date_text
                        break
        
        return {
            'headline': headline,
            'url': full_url,
            'article_type': article_type,
            'published_date': published_date,
            'scraped_date': datetime.now(timezone.utc).isoformat(),
            'hash_id': generate_hash_id(full_url.split('?')[0])
        }
    except:
        return None

def scrape_articles_from_page(driver, existing_ids: set, page_num: int = 1) -> List[Dict]:
    """Scrape articles from listing page"""
    try:
        url = f"{START_URL}?page={page_num-1}" if page_num > 1 else START_URL
        logging.info(f"Scraping page {page_num}: {url}")
        
        driver.get(url)
        WebDriverWait(driver, PAGE_LOAD_TIMEOUT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(3)
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        # Find article links
        article_links = soup.select('a[href*="news-and-publications"]')
        
        valid_articles = []
        processed_urls = set()
        
        for link in article_links:
            href = link.get('href')
            if (href and 
                not any(param in href for param in ['?page=', '?industry=', '?tags=']) and
                href.count('/') >= 2):
                
                metadata = extract_article_metadata(link, BASE_URL)
                if (metadata and 
                    metadata['url'] not in processed_urls and 
                    metadata['hash_id'] not in existing_ids):
                    valid_articles.append(metadata)
                    processed_urls.add(metadata['url'])
        
        logging.info(f"Found {len(valid_articles)} new articles on page {page_num}")
        return valid_articles
        
    except Exception as e:
        logging.error(f"Error scraping page {page_num}: {e}")
        return []

def check_pagination(driver) -> bool:
    try:
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        next_selectors = ['.pagination__next a', 'a[rel="next"]', '.next a']
        return any(soup.select_one(selector) for selector in next_selectors)
    except:
        return False

def load_existing_hash_ids() -> set:
    if os.path.exists(JSON_PATH):
        try:
            with open(JSON_PATH, 'r', encoding='utf-8') as f:
                articles = json.load(f)
                return {article.get("hash_id") for article in articles if article.get("hash_id")}
        except:
            return set()
    return set()

def save_articles(articles: List[Dict]):
    """Save articles with clean JSON structure"""
    try:
        # Load existing
        all_articles = []
        if os.path.exists(JSON_PATH):
            try:
                with open(JSON_PATH, 'r', encoding='utf-8') as f:
                    all_articles = json.load(f)
            except:
                all_articles = []
        
        # Add new articles
        all_articles.extend(articles)
        
        # Save with clean formatting
        with open(JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(all_articles, f, indent=2, ensure_ascii=False)
        
        # Stats
        total_tokens = sum(art.get('total_token_count', 0) for art in articles)
        total_attachments = sum(art.get('attachments_count', 0) for art in articles)
        total_linked = sum(art.get('linked_pages_count', 0) for art in articles)
        
        logging.info(f"Saved {len(articles)} articles")
        logging.info(f"Total tokens: {total_tokens:,}")
        logging.info(f"Total attachments: {total_attachments}")
        logging.info(f"Total linked pages: {total_linked}")
        logging.info(f"Output: {JSON_PATH}")
        
        return True
    except Exception as e:
        logging.error(f"Error saving: {e}")
        return False

def main():
    setup_logging()
    logging.info("="*60)
    logging.info("IMPROVED APRA SCRAPER - COMPREHENSIVE CONTENT EXTRACTION")
    logging.info("="*60)
    
    try:
        driver = setup_stealth_driver()
        session = create_session()
        
        try:
            # Simulate human browsing
            driver.get(BASE_URL)
            human_delay()
            driver.get(START_URL)
            human_delay()
            
            existing_ids = load_existing_hash_ids()
            all_articles = []
            current_page = 1
            
            # Scrape article listings
            while True:
                page_articles = scrape_articles_from_page(driver, existing_ids, current_page)
                
                if not page_articles:
                    if not check_pagination(driver):
                        break
                    current_page += 1
                    continue
                else:
                    all_articles.extend(page_articles)
                
                if MAX_PAGES and current_page >= MAX_PAGES:
                    break
                
                if not check_pagination(driver):
                    break
                    
                current_page += 1
                human_delay()
            
            if all_articles:
                logging.info(f"Processing {len(all_articles)} articles with full content extraction...")
                
                # Process each article with comprehensive content extraction
                enriched_articles = []
                for i, article in enumerate(all_articles):
                    logging.info(f"Processing {i+1}/{len(all_articles)}: {article['headline'][:50]}...")
                    
                    # Get comprehensive content
                    content_result = process_article_with_links(article['url'], driver, session)
                    
                    if 'error' not in content_result:
                        # Clean JSON structure
                        enriched_article = {
                            "hash_id": article['hash_id'],
                            "headline": content_result['headline'],
                            "url": article['url'],
                            "article_type": content_result['article_type'],
                            "published_date": content_result['published_date'],
                            "scraped_date": article['scraped_date'],
                            
                            # Content (includes PDF text)
                            "content": content_result['main_content'],
                            
                            # Linked pages content (without breadcrumbs)
                            "linked_pages": content_result['linked_pages'],
                            
                            # All attachments from main + linked pages (PDF and Excel only)
                            "attachments": content_result['all_attachments'],
                            
                            # Token counts for LLM analysis
                            "total_token_count": content_result['total_token_count'],
                            "main_token_count": content_result['main_token_count'],
                            "linked_token_count": content_result['linked_token_count'],
                            
                            # Summary stats
                            "linked_pages_count": content_result['linked_pages_count'],
                            "attachments_count": content_result['attachments_count'],
                            
                            # LLM readiness
                            "llm_ready": content_result['total_token_count'] > 50
                        }
                        
                        enriched_articles.append(enriched_article)
                        logging.info(f"✓ Extracted: {enriched_article['linked_pages_count']} linked pages, "
                                   f"{enriched_article['attachments_count']} attachments, "
                                   f"{enriched_article['total_token_count']:,} tokens")
                    else:
                        # Failed article
                        enriched_articles.append({
                            "hash_id": article['hash_id'],
                            "headline": article['headline'],
                            "url": article['url'],
                            "article_type": article['article_type'],
                            "published_date": article['published_date'],
                            "scraped_date": article['scraped_date'],
                            "content": "",
                            "linked_pages": [],
                            "attachments": [],
                            "total_token_count": 0,
                            "main_token_count": 0,
                            "linked_token_count": 0,
                            "linked_pages_count": 0,
                            "attachments_count": 0,
                            "llm_ready": False,
                            "error": content_result.get('error', 'Unknown error')
                        })
                    
                    human_delay()
                
                if save_articles(enriched_articles):
                    logging.info("✅ SUCCESS: Improved APRA scraping completed!")
                    return True
                else:
                    return False
            else:
                logging.info("No new articles found")
                return True
                
        finally:
            driver.quit()
            session.close()
            
    except Exception as e:
        logging.error(f"Critical error: {e}")
        return False
    
    finally:
        logging.info("="*60)
        logging.info("IMPROVED SCRAPER FINISHED")
        logging.info("="*60)

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)