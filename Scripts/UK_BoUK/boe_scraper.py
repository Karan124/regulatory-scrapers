#!/usr/bin/env python3
"""
Bank of England News & Publications Scraper

A production-ready scraper for all News & Publications from the Bank of England site.
Handles JavaScript rendering, PDF extraction, deduplication, and structured data output.

Usage:
    python boe_scraper.py                    # Incremental run (default)
    python boe_scraper.py --full-refresh     # Full refresh
    python boe_scraper.py --max-pages 5      # Limit to 5 pages
"""

import asyncio
import hashlib
import json
import logging
import os
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Any, Set
from urllib.parse import urljoin, urlparse
import argparse
import csv
import tempfile

# Core dependencies
import requests
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, Page, Browser
import pandas as pd

# PDF and document processing
try:
    import pdfplumber
    import tabula
except ImportError:
    print("Warning: PDF processing libraries not installed. Run: pip install pdfplumber tabula-py")
    pdfplumber = None
    tabula = None

try:
    import openpyxl
except ImportError:
    print("Warning: Excel processing library not installed. Run: pip install openpyxl")
    openpyxl = None

# Rate limiting and retries
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

# Configuration
DATA_DIR = "data"
OUTPUT_FILE = "bouk_news.json"
DELTA_FILE = "bouk_news_delta.jsonl"
CSV_INDEX_FILE = "bouk_news_index.csv"
ATTACHMENTS_DIR = "attachments"
SCREENSHOTS_DIR = "screenshots"

# Default settings
DEFAULT_MAX_PAGES = "ALL"
DEFAULT_REQUESTS_PER_MINUTE = 30
DEFAULT_USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"

# URLs
BASE_URL = "https://www.bankofengland.co.uk"
NEWS_URL = f"{BASE_URL}/news"

# Marketing/social media domains to exclude from embedded links
MARKETING_DOMAINS = {
    'facebook.com', 'instagram.com', 'twitter.com', 'x.com', 'linkedin.com',
    'youtube.com', 'tiktok.com', 'flickr.com', 'threads.net', 'snapchat.com'
}

class BOEScraperConfig:
    """Configuration class for the scraper"""
    def __init__(self):
        self.data_dir = DATA_DIR
        self.output_file = OUTPUT_FILE
        self.max_pages = DEFAULT_MAX_PAGES
        self.requests_per_minute = DEFAULT_REQUESTS_PER_MINUTE
        self.user_agent = DEFAULT_USER_AGENT
        self.update_existing = True
        self.save_pdfs = False
        self.save_screenshots = False
        self.playwright_headless = True
        self.proxy_list = []

class RateLimiter:
    """Simple rate limiter"""
    def __init__(self, requests_per_minute: int):
        self.requests_per_minute = requests_per_minute
        self.min_interval = 60.0 / requests_per_minute
        self.last_request = 0
    
    async def wait(self):
        """Wait if necessary to respect rate limit"""
        now = time.time()
        elapsed = now - self.last_request
        if elapsed < self.min_interval:
            sleep_time = self.min_interval - elapsed
            await asyncio.sleep(sleep_time)
        self.last_request = time.time()

class BOEScraper:
    """Main scraper class for Bank of England news and publications"""
    
    def __init__(self, config: BOEScraperConfig):
        self.config = config
        self.rate_limiter = RateLimiter(config.requests_per_minute)
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': config.user_agent})
        
        # Setup directories
        self.data_path = Path(config.data_dir)
        self.data_path.mkdir(exist_ok=True)
        (self.data_path / ATTACHMENTS_DIR).mkdir(exist_ok=True)
        (self.data_path / SCREENSHOTS_DIR).mkdir(exist_ok=True)
        
        # Load existing data for deduplication
        self.existing_ids = self._load_existing_ids()
        self.new_articles = []
        
        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(self.data_path / 'scraper.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def _load_existing_ids(self) -> Set[str]:
        """Load existing article IDs for deduplication"""
        output_path = self.data_path / self.config.output_file
        if not output_path.exists():
            return set()
        
        try:
            with open(output_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {article['id'] for article in data}
        except (json.JSONDecodeError, KeyError) as e:
            self.logger.warning(f"Could not load existing IDs: {e}")
            return set()

    def _generate_article_id(self, url: str, headline: str, published_date: str, body_preview: str = "") -> str:
        """Generate stable unique ID for article"""
        # Prefer URL-based ID, fallback to content hash
        if url and url.startswith('http'):
            url_hash = hashlib.sha256(url.encode()).hexdigest()[:16]
            return f"url:{url_hash}"
        
        # Content-based hash
        content = f"{headline}|{published_date}|{body_preview[:1000]}"
        content_hash = hashlib.sha256(content.encode()).hexdigest()[:16]
        return f"sha256:{content_hash}"

    async def _setup_browser_session(self, page: Page) -> None:
        """Setup browser session with cookies and realistic behavior"""
        try:
            # Navigate to homepage first
            await page.goto(BASE_URL, wait_until='networkidle')
            await asyncio.sleep(2)
            
            # Accept cookies if banner present
            try:
                # Common cookie banner selectors
                cookie_selectors = [
                    'button[id*="cookie"]',
                    'button[class*="cookie"]',
                    'button:has-text("Accept")',
                    'button:has-text("Agree")',
                    '[data-accept="cookies"]'
                ]
                
                for selector in cookie_selectors:
                    try:
                        await page.click(selector, timeout=2000)
                        self.logger.info("Accepted cookies")
                        await asyncio.sleep(1)
                        break
                    except:
                        continue
                        
            except Exception as e:
                self.logger.debug(f"No cookie banner found or couldn't click: {e}")
            
            # Navigate around to build session
            await page.goto(f"{BASE_URL}/about", wait_until='networkidle')
            await asyncio.sleep(1)
            
        except Exception as e:
            self.logger.warning(f"Could not complete browser session setup: {e}")

    async def _scrape_news_listing(self, page: Page) -> List[Dict[str, Any]]:
        """Scrape the news listing pages"""
        articles = []
        page_num = 1
        max_pages = self.config.max_pages
        
        # Navigate to news page
        await page.goto(NEWS_URL, wait_until='networkidle')
        await asyncio.sleep(3)
        
        while True:
            self.logger.info(f"Scraping page {page_num}")
            
            # Wait for results to load
            try:
                await page.wait_for_selector('#SearchResults', timeout=10000)
            except:
                self.logger.error("Could not find search results")
                break
            
            # Extract articles from current page
            page_articles = await self._extract_articles_from_page(page)
            articles.extend(page_articles)
            
            self.logger.info(f"Found {len(page_articles)} articles on page {page_num}")
            
            # Check if we should continue
            if max_pages != "ALL" and page_num >= int(max_pages):
                self.logger.info(f"Reached max pages limit: {max_pages}")
                break
            
            # Try to go to next page
            try:
                next_button = await page.query_selector('a.list-pagination__link--next')
                if not next_button:
                    self.logger.info("No next page button found")
                    break
                
                # Click next page
                await next_button.click()
                await page.wait_for_selector('#SearchResults', timeout=10000)
                await asyncio.sleep(2)
                page_num += 1
                
            except Exception as e:
                self.logger.info(f"No more pages available: {e}")
                break
        
        return articles

    async def _extract_articles_from_page(self, page: Page) -> List[Dict[str, Any]]:
        """Extract article metadata from a listing page"""
        articles = []
        
        # Get page content
        content = await page.content()
        soup = BeautifulSoup(content, 'html.parser')
        
        # Find article containers
        article_containers = soup.find_all('div', class_='col3')
        
        for container in article_containers:
            try:
                article_link = container.find('a', class_='release')
                if not article_link:
                    continue
                
                # Extract basic metadata
                url = urljoin(BASE_URL, article_link.get('href', ''))
                
                # Get title
                title_elem = article_link.find('h3')
                headline = title_elem.get_text(strip=True) if title_elem else ""
                
                # Get category/theme
                tag_elem = container.find('div', class_='release-tag')
                theme = tag_elem.get_text(strip=True) if tag_elem else ""
                
                # Get date
                date_elem = container.find('time', class_='release-date')
                published_date = ""
                if date_elem:
                    datetime_attr = date_elem.get('datetime')
                    if datetime_attr:
                        # Convert to ISO format
                        try:
                            dt = datetime.fromisoformat(datetime_attr).replace(tzinfo=timezone.utc)
                            published_date = dt.isoformat()
                        except:
                            published_date = datetime_attr
                
                # Generate ID and check for duplicates
                article_id = self._generate_article_id(url, headline, published_date)
                
                if article_id in self.existing_ids:
                    self.logger.info(f"Skipping duplicate article: {headline}")
                    continue
                
                article_data = {
                    'id': article_id,
                    'url': url,
                    'headline': headline,
                    'theme': theme,
                    'published_date': published_date,
                    'scraped_date': datetime.now(timezone.utc).isoformat(),
                    'source_site': 'bankofengland.co.uk'
                }
                
                articles.append(article_data)
                
            except Exception as e:
                self.logger.error(f"Error extracting article metadata: {e}")
                continue
        
        return articles

    async def _scrape_article_content(self, page: Page, article: Dict[str, Any]) -> Dict[str, Any]:
        """Scrape full content of an individual article"""
        await self.rate_limiter.wait()
        
        try:
            self.logger.info(f"Scraping article: {article['headline']}")
            await page.goto(article['url'], wait_until='networkidle')
            await asyncio.sleep(2)
            
            content = await page.content()
            soup = BeautifulSoup(content, 'html.parser')
            
            # Extract summary/description
            summary = ""
            summary_elem = soup.find('div', class_='page-description')
            if summary_elem:
                summary = summary_elem.get_text(strip=True)
            
            # Extract main content
            body_text = ""
            main_content = soup.find('div', class_='container-publication') or soup.find('main')
            if main_content:
                # Remove navigation and other non-content elements
                for elem in main_content.find_all(['nav', 'script', 'style', 'button', 'form']):
                    elem.decompose()
                
                body_text = main_content.get_text(separator='\n', strip=True)
                # Clean up whitespace
                body_text = re.sub(r'\n\s*\n\s*\n', '\n\n', body_text)
                body_text = re.sub(r'\s+', ' ', body_text)
            
            # Extract authors with improved filtering
            authors = []
            body_lower = body_text.lower()
            
            # Common author patterns for Bank of England content
            author_patterns = [
                r'speech by ([A-Z][a-z]+ [A-Z][a-z]+)',  # "speech by John Smith"
                r'written by ([A-Z][a-z]+ [A-Z][a-z]+)',  # "written by Jane Doe"
                r'by ([A-Z][a-z]+ [A-Z][a-z]+),?\s+(?:Deputy Governor|Governor|Executive Director)',  # "by John Smith, Governor"
                r'(?:Governor|Deputy Governor|Executive Director)\s+([A-Z][a-z]+ [A-Z][a-z]+)',  # "Governor John Smith"
            ]
            
            for pattern in author_patterns:
                matches = re.findall(pattern, body_text, re.IGNORECASE)
                for match in matches:
                    # Clean up the author name
                    author = match.strip()
                    # Only add if it looks like a real name (avoid false positives)
                    if (len(author.split()) == 2 and 
                        not any(word in author.lower() for word in ['margin', 'sample', 'survey', 'between', 'since', 'note', 'change']) and
                        all(len(word) > 1 for word in author.split())):
                        authors.append(author)
            
            # Remove duplicates and limit to reasonable number
            authors = list(set(authors))[:5]
            
            # Extract main image
            main_image_url = ""
            img_elem = soup.find('img')
            if img_elem and img_elem.get('src'):
                main_image_url = urljoin(BASE_URL, img_elem['src'])
            
            # Extract embedded links (exclude marketing/social media)
            embedded_links = []
            
            for link in soup.find_all('a', href=True):
                href = link['href']
                if href.startswith('http') and not href.startswith(BASE_URL):
                    # Skip marketing/social media links
                    link_domain = urlparse(href).netloc.lower()
                    if not any(domain in link_domain for domain in MARKETING_DOMAINS):
                        # Only include if it's in main content area, not footer/nav
                        parent_classes = ' '.join(link.parent.get('class', []) if link.parent else [])
                        parent_text = str(link.parent) if link.parent else ""
                        
                        if not any(skip_term in parent_classes.lower() + parent_text.lower() 
                                 for skip_term in ['footer', 'navigation', 'nav', 'menu', 'social']):
                            embedded_links.append(href)
            
            # Extract tables with NaN handling
            tables_extracted = []
            for i, table in enumerate(soup.find_all('table')):
                try:
                    df = pd.read_html(str(table))[0]
                    
                    # Clean NaN values and convert to JSON-safe format
                    df = df.fillna('')  # Replace NaN with empty strings
                    headers = [str(col) for col in df.columns.tolist()]
                    rows = []
                    
                    for _, row in df.iterrows():
                        clean_row = []
                        for cell in row:
                            if pd.isna(cell) or str(cell).lower() == 'nan':
                                clean_row.append('')
                            else:
                                clean_row.append(str(cell))
                        rows.append(clean_row)
                    
                    table_data = {
                        'table_id': f't{i+1}',
                        'headers': headers,
                        'rows': rows
                    }
                    tables_extracted.append(table_data)
                except Exception as e:
                    self.logger.debug(f"Could not parse table {i}: {e}")
            
            # Find PDF and other attachments (don't save files)
            attachments = []
            for link in soup.find_all('a', href=True):
                href = link['href']
                if any(href.lower().endswith(ext) for ext in ['.pdf', '.xlsx', '.xls', '.csv']):
                    attachment_url = urljoin(BASE_URL, href)
                    filename = os.path.basename(urlparse(attachment_url).path)
                    
                    attachment_data = {
                        'filename': filename,
                        'url': attachment_url,
                        'mime_type': self._get_mime_type(filename),
                        'extracted_text_key': None
                    }
                    
                    # Download and extract content (but don't save files locally)
                    extracted_text = await self._download_and_extract_attachment(
                        attachment_url, article['id'], filename, save_file=False
                    )
                    if extracted_text:
                        # Add extracted text directly to body instead of saving file reference
                        body_text += f"\n\n--- Content from {filename} ---\n{extracted_text}"
                    
                    attachments.append(attachment_data)
            
            # Look for charts/visualizations
            charts_data = []
            for iframe in soup.find_all('iframe'):
                src = iframe.get('src', '')
                if 'tableau' in src.lower():
                    chart_data = {
                        'chart_type': 'tableau',
                        'source': src,
                        'data_csv_url': self._try_extract_tableau_data_url(src)
                    }
                    charts_data.append(chart_data)
            
            # Update article with extracted content
            article.update({
                'summary': summary,
                'body_text': body_text,
                'authors': authors,
                'main_image_url': main_image_url,
                'embedded_links': list(set(embedded_links)),  # Remove duplicates
                'attachments': attachments,
                'tables_extracted': tables_extracted,
                'charts_data': charts_data
            })
            
            return article
            
        except Exception as e:
            self.logger.error(f"Error scraping article {article['url']}: {e}")
            article['scrape_error'] = str(e)
            return article

    def _get_mime_type(self, filename: str) -> str:
        """Get MIME type from filename"""
        ext = os.path.splitext(filename)[1].lower()
        mime_types = {
            '.pdf': 'application/pdf',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.csv': 'text/csv'
        }
        return mime_types.get(ext, 'application/octet-stream')

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
    async def _download_and_extract_attachment(self, url: str, article_id: str, filename: str, save_file: bool = False) -> Optional[str]:
        """Download and extract text from attachments"""
        try:
            await self.rate_limiter.wait()
            
            # Download file
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            # Only save file if explicitly requested (disabled by default)
            if save_file:
                attachment_dir = self.data_path / ATTACHMENTS_DIR / article_id
                attachment_dir.mkdir(exist_ok=True, parents=True)
                
                file_path = attachment_dir / filename
                with open(file_path, 'wb') as f:
                    f.write(response.content)
            else:
                # Save to temporary file for processing only
                with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp_file:
                    tmp_file.write(response.content)
                    file_path = Path(tmp_file.name)
            
            # Extract text based on file type
            extracted_text = ""
            
            if filename.lower().endswith('.pdf') and pdfplumber:
                extracted_text = self._extract_pdf_text(file_path)
            elif filename.lower().endswith(('.xlsx', '.xls')) and openpyxl:
                extracted_text = self._extract_excel_text(file_path)
            elif filename.lower().endswith('.csv'):
                extracted_text = self._extract_csv_text(file_path)
            
            # Clean up temporary file if not saving permanently
            if not save_file and file_path.exists():
                try:
                    file_path.unlink()
                except:
                    pass  # Ignore cleanup errors
            
            return extracted_text
            
        except Exception as e:
            self.logger.error(f"Error downloading/extracting {url}: {e}")
            return None

    def _extract_pdf_text(self, file_path: Path) -> str:
        """Extract text from PDF"""
        if not pdfplumber:
            return ""
        
        try:
            text_content = []
            
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    # Extract text
                    text = page.extract_text()
                    if text:
                        text_content.append(text)
                    
                    # Extract tables
                    tables = page.extract_tables()
                    for table in tables:
                        if table:
                            table_text = "\n".join(["\t".join([cell or "" for cell in row]) for row in table])
                            text_content.append(f"\n[TABLE]\n{table_text}\n[/TABLE]\n")
            
            return "\n\n".join(text_content)
            
        except Exception as e:
            self.logger.error(f"Error extracting PDF text from {file_path}: {e}")
            return ""

    def _extract_excel_text(self, file_path: Path) -> str:
        """Extract text from Excel files - improved to handle all sheets properly"""
        try:
            text_content = []
            
            # Use openpyxl for .xlsx files to get all data including hidden sheets
            if file_path.suffix.lower() == '.xlsx':
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_content.append(f"[SHEET: {sheet_name}]")
                    
                    # Get all data from the sheet
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):  # Skip completely empty rows
                            clean_row = []
                            for cell in row:
                                if cell is None or pd.isna(cell):
                                    clean_row.append('')
                                else:
                                    clean_row.append(str(cell))
                            text_content.append('\t'.join(clean_row))
                    
                    text_content.append("")  # Empty line between sheets
                
                wb.close()
            else:
                # For .xls files, use pandas
                xl_file = pd.ExcelFile(file_path)
                
                for sheet_name in xl_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    df = df.fillna('')  # Replace NaN with empty strings
                    text_content.append(f"[SHEET: {sheet_name}]")
                    text_content.append(df.to_string(index=False))
                    text_content.append("")
            
            return "\n".join(text_content)
            
        except Exception as e:
            self.logger.error(f"Error extracting Excel text from {file_path}: {e}")
            return ""

    def _extract_csv_text(self, file_path: Path) -> str:
        """Extract text from CSV files"""
        try:
            df = pd.read_csv(file_path)
            df = df.fillna('')  # Replace NaN with empty strings
            return df.to_string(index=False)
        except Exception as e:
            self.logger.error(f"Error extracting CSV text from {file_path}: {e}")
            return ""

    def _try_extract_tableau_data_url(self, iframe_src: str) -> Optional[str]:
        """Try to extract data URL from Tableau iframe"""
        try:
            # Basic pattern matching for Tableau public data URLs
            if 'public.tableau.com' in iframe_src:
                # Try to construct CSV download URL
                if '/views/' in iframe_src:
                    base_url = iframe_src.split('?')[0]
                    data_url = f"{base_url}.csv"
                    return data_url
        except Exception:
            pass
        return None

    async def run(self) -> None:
        """Main scraper execution"""
        self.logger.info("Starting Bank of England scraper")
        
        async with async_playwright() as p:
            # Launch browser
            browser = await p.chromium.launch(
                headless=self.config.playwright_headless,
                args=['--no-sandbox', '--disable-blink-features=AutomationControlled']
            )
            
            # Create context with stealth settings
            context = await browser.new_context(
                user_agent=self.config.user_agent,
                viewport={'width': 1920, 'height': 1080},
                extra_http_headers={'Accept-Language': 'en-US,en;q=0.9'}
            )
            
            # Add stealth script
            await context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => false});
                Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
                Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
            """)
            
            page = await context.new_page()
            
            try:
                # Setup browser session
                await self._setup_browser_session(page)
                
                # Scrape article listings
                articles = await self._scrape_news_listing(page)
                self.logger.info(f"Found {len(articles)} articles to process")
                
                # Scrape full content for each article
                for i, article in enumerate(articles):
                    try:
                        self.logger.info(f"Processing article {i+1}/{len(articles)}: {article['headline']}")
                        full_article = await self._scrape_article_content(page, article)
                        self.new_articles.append(full_article)
                        
                        # Save progress periodically
                        if (i + 1) % 10 == 0:
                            self._save_progress()
                            
                    except Exception as e:
                        self.logger.error(f"Error processing article {i+1}: {e}")
                        continue
                
                # Final save
                self._save_results()
                
            finally:
                await browser.close()

    def _save_progress(self) -> None:
        """Save progress incrementally"""
        if not self.new_articles:
            return
        
        # Save delta file
        delta_path = self.data_path / DELTA_FILE
        with open(delta_path, 'a', encoding='utf-8') as f:
            for article in self.new_articles:
                f.write(json.dumps(article, ensure_ascii=False) + '\n')

    def _save_results(self) -> None:
        """Save final results"""
        self.logger.info(f"Saving {len(self.new_articles)} new articles")
        
        # Load existing data
        output_path = self.data_path / self.config.output_file
        all_articles = []
        
        if output_path.exists():
            try:
                with open(output_path, 'r', encoding='utf-8') as f:
                    all_articles = json.load(f)
            except json.JSONDecodeError:
                self.logger.warning("Could not load existing data, starting fresh")
        
        # Add new articles
        all_articles.extend(self.new_articles)
        
        # Save main JSON file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(all_articles, f, ensure_ascii=False, indent=2)
        
        # Save delta JSONL
        delta_path = self.data_path / DELTA_FILE
        with open(delta_path, 'w', encoding='utf-8') as f:
            for article in self.new_articles:
                f.write(json.dumps(article, ensure_ascii=False) + '\n')
        
        # Save CSV index
        self._save_csv_index(all_articles)
        
        self.logger.info(f"Results saved to {output_path}")
        self.logger.info(f"Total articles: {len(all_articles)}")

    def _save_csv_index(self, articles: List[Dict[str, Any]]) -> None:
        """Save CSV index of articles"""
        csv_path = self.data_path / CSV_INDEX_FILE
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['id', 'headline', 'published_date', 'url', 'scraped_date', 'theme'])
            
            for article in articles:
                writer.writerow([
                    article.get('id', ''),
                    article.get('headline', ''),
                    article.get('published_date', ''),
                    article.get('url', ''),
                    article.get('scraped_date', ''),
                    article.get('theme', '')
                ])

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="Bank of England News Scraper")
    parser.add_argument('--full-refresh', action='store_true', help='Perform full refresh')
    parser.add_argument('--max-pages', type=str, default=DEFAULT_MAX_PAGES, help='Maximum pages to scrape')
    parser.add_argument('--save-pdfs', action='store_true', help='Save PDF files locally')
    parser.add_argument('--save-screenshots', action='store_true', help='Save screenshots')
    parser.add_argument('--headless', action='store_true', default=True, help='Run browser in headless mode')
    parser.add_argument('--requests-per-minute', type=int, default=DEFAULT_REQUESTS_PER_MINUTE, help='Rate limit')
    
    args = parser.parse_args()
    
    # Setup configuration
    config = BOEScraperConfig()
    config.max_pages = args.max_pages
    config.save_pdfs = args.save_pdfs
    config.save_screenshots = args.save_screenshots
    config.playwright_headless = args.headless
    config.requests_per_minute = args.requests_per_minute
    
    # If full refresh, clear existing data
    if args.full_refresh:
        output_path = Path(config.data_dir) / config.output_file
        if output_path.exists():
            output_path.unlink()
        print("Full refresh mode: cleared existing data")
    
    # Run scraper
    scraper = BOEScraper(config)
    asyncio.run(scraper.run())

if __name__ == "__main__":
    main()