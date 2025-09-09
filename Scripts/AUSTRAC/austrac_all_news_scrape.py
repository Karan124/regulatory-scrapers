#!/usr/bin/env python3
"""
Enhanced AUSTRAC Updates Scraper - Complete Working Version
- LLM-friendly content extraction with structured formatting
- PDF and Excel/CSV document extraction with full text content
- JSON-only output optimized for LLM analysis
- Enhanced content cleaning and structuring
- Daily vs Initial run differentiation
"""

import os
import json
import time
import hashlib
import logging
import re
import random
import io
import subprocess
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Optional, Any
from pathlib import Path
from urllib.parse import urljoin, urlparse
import requests
import signal
import sys

# Core scraping libraries
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from bs4 import BeautifulSoup
import urllib3

# Document processing libraries
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import pymupdf as fitz
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# Disable warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ----------------------------
# Enhanced Logging Setup
# ----------------------------
os.makedirs('data', exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s [%(funcName)s:%(lineno)d]: %(message)s',
    handlers=[
        logging.FileHandler('data/austrac_updates_enhanced.log', mode='a', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Log available document processing libraries
logger.info(f"Document processing libraries: PyPDF2={HAS_PYPDF2}, PyMuPDF={HAS_PYMUPDF}, pdfplumber={HAS_PDFPLUMBER}, openpyxl={HAS_OPENPYXL}, pandas={HAS_PANDAS}")

# ----------------------------
# Configuration
# ----------------------------
DATA_DIR = Path("data")
JSON_PATH = DATA_DIR / "austrac_news.json"

BASE_URL = "https://www.austrac.gov.au"
TARGET_URL = f"{BASE_URL}/business/updates"
PAGE_LOAD_TIMEOUT = 30
ARTICLE_TIMEOUT = 20
DELAY_BETWEEN_REQUESTS = 3
MAX_RETRIES = 3
MAX_PAGES_INITIAL = 2  # For comprehensive initial scrapes
MAX_PAGES_DAILY = 1     # For daily runs

# Global variables for graceful shutdown
shutdown_requested = False

def signal_handler(signum, frame):
    global shutdown_requested
    logger.info(f"Received signal {signum}. Initiating graceful shutdown...")
    shutdown_requested = True

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

class DocumentProcessor:
    """Enhanced document processor for PDFs, Excel, and CSV files"""
    
    @staticmethod
    def clean_text_for_llm(text: str) -> str:
        """Clean and structure text content for optimal LLM processing"""
        if not text:
            return ""
        
        # Remove excessive whitespace while preserving paragraph structure
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)
        text = re.sub(r'[ \t]+', ' ', text)
        text = re.sub(r'^\s+|\s+$', '', text, flags=re.MULTILINE)
        
        # Remove page numbers and headers/footers
        text = re.sub(r'\n\s*\d+\s*\n', '\n', text)
        text = re.sub(r'\n\s*Page \d+ of \d+\s*\n', '\n', text, flags=re.IGNORECASE)
        
        # Clean up common document artifacts
        text = re.sub(r'[^\w\s\.\,\;\:\!\?\-\(\)\[\]\{\}\"\'\/\\\@\#\$\%\&\*\+\=\<\>\~\`\|\n]', '', text)
        
        # Ensure proper sentence spacing
        text = re.sub(r'\.([A-Z])', r'. \1', text)
        text = re.sub(r'\?([A-Z])', r'? \1', text)
        text = re.sub(r'\!([A-Z])', r'! \1', text)
        
        return text.strip()
    
    @staticmethod
    def extract_pdf_content(pdf_url: str) -> Dict[str, Any]:
        """Extract comprehensive content from PDF with multiple methods"""
        result = {
            'success': False,
            'text_content': '',
            'metadata': {},
            'extraction_method': '',
            'page_count': 0,
            'file_size_mb': 0,
            'error': None
        }
        
        try:
            logger.info(f"Downloading PDF: {pdf_url}")
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            }
            
            response = requests.get(pdf_url, headers=headers, timeout=60, verify=False)
            if response.status_code != 200:
                result['error'] = f"HTTP {response.status_code}"
                return result
            
            pdf_content = response.content
            result['file_size_mb'] = round(len(pdf_content) / 1024 / 1024, 2)
            
            extracted_text = ""
            
            # Method 1: Try PyMuPDF (best for comprehensive extraction)
            if HAS_PYMUPDF and not extracted_text:
                try:
                    pdf_document = fitz.open(stream=pdf_content, filetype="pdf")
                    result['page_count'] = pdf_document.page_count
                    
                    # Extract metadata
                    metadata = pdf_document.metadata
                    result['metadata'] = {
                        'title': metadata.get('title', ''),
                        'author': metadata.get('author', ''),
                        'subject': metadata.get('subject', ''),
                        'creator': metadata.get('creator', ''),
                        'creation_date': metadata.get('creationDate', ''),
                        'modification_date': metadata.get('modDate', '')
                    }
                    
                    text_parts = []
                    for page_num in range(pdf_document.page_count):
                        page = pdf_document[page_num]
                        
                        # Extract text blocks for better structure
                        blocks = page.get_text("blocks")
                        page_text_parts = []
                        
                        for block in blocks:
                            if len(block) > 4 and block[4].strip():
                                page_text_parts.append(block[4].strip())
                        
                        if page_text_parts:
                            page_text = "\n".join(page_text_parts)
                            text_parts.append(f"\n--- PAGE {page_num + 1} ---\n{page_text}")
                    
                    pdf_document.close()
                    
                    if text_parts:
                        extracted_text = "\n\n".join(text_parts)
                        result['extraction_method'] = 'PyMuPDF'
                        logger.info(f"PyMuPDF extracted {len(extracted_text)} characters from {result['page_count']} pages")
                
                except Exception as e:
                    logger.warning(f"PyMuPDF extraction failed: {e}")
            
            # Method 2: Try pdfplumber (excellent for tables and structured content)
            if HAS_PDFPLUMBER and not extracted_text:
                try:
                    with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
                        result['page_count'] = len(pdf.pages)
                        text_parts = []
                        
                        for page_num, page in enumerate(pdf.pages):
                            # Extract regular text
                            page_text = page.extract_text()
                            if page_text and page_text.strip():
                                text_parts.append(f"\n--- PAGE {page_num + 1} ---\n{page_text.strip()}")
                            
                            # Extract tables separately
                            tables = page.extract_tables()
                            if tables:
                                for table_num, table in enumerate(tables):
                                    if table:
                                        table_text = "\n".join([
                                            " | ".join([str(cell) if cell else "" for cell in row])
                                            for row in table if row
                                        ])
                                        if table_text.strip():
                                            text_parts.append(f"\n--- TABLE {table_num + 1} (PAGE {page_num + 1}) ---\n{table_text}")
                        
                        if text_parts:
                            extracted_text = "\n\n".join(text_parts)
                            result['extraction_method'] = 'pdfplumber'
                            logger.info(f"pdfplumber extracted {len(extracted_text)} characters from {result['page_count']} pages")
                
                except Exception as e:
                    logger.warning(f"pdfplumber extraction failed: {e}")
            
            # Method 3: Fallback to PyPDF2
            if HAS_PYPDF2 and not extracted_text:
                try:
                    pdf_file = io.BytesIO(pdf_content)
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    result['page_count'] = len(pdf_reader.pages)
                    
                    text_parts = []
                    for page_num, page in enumerate(pdf_reader.pages):
                        page_text = page.extract_text()
                        if page_text and page_text.strip():
                            text_parts.append(f"\n--- PAGE {page_num + 1} ---\n{page_text.strip()}")
                    
                    if text_parts:
                        extracted_text = "\n\n".join(text_parts)
                        result['extraction_method'] = 'PyPDF2'
                        logger.info(f"PyPDF2 extracted {len(extracted_text)} characters from {result['page_count']} pages")
                
                except Exception as e:
                    logger.warning(f"PyPDF2 extraction failed: {e}")
            
            if extracted_text:
                # Clean and structure text for LLM
                cleaned_text = DocumentProcessor.clean_text_for_llm(extracted_text)
                
                # Add extraction metadata
                metadata_header = f"""--- PDF DOCUMENT METADATA ---
Source URL: {pdf_url}
Extraction Date: {datetime.now().isoformat()}
File Size: {result['file_size_mb']} MB
Page Count: {result['page_count']}
Extraction Method: {result['extraction_method']}
Content Length: {len(cleaned_text)} characters
--- END METADATA ---

"""
                
                result['text_content'] = metadata_header + cleaned_text
                result['success'] = True
                
                logger.info(f"Successfully extracted PDF content: {len(result['text_content'])} characters")
            else:
                result['error'] = "No text could be extracted"
                logger.warning(f"No text extracted from PDF: {pdf_url}")
            
            return result
            
        except Exception as e:
            result['error'] = str(e)
            logger.error(f"Error extracting PDF {pdf_url}: {e}")
            return result
    
    @staticmethod
    def extract_excel_content(excel_url: str) -> Dict[str, Any]:
        """Extract content from Excel/CSV files"""
        result = {
            'success': False,
            'text_content': '',
            'sheet_data': {},
            'metadata': {},
            'file_size_mb': 0,
            'error': None
        }
        
        try:
            logger.info(f"Downloading Excel/CSV: {excel_url}")
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.get(excel_url, headers=headers, timeout=60, verify=False)
            if response.status_code != 200:
                result['error'] = f"HTTP {response.status_code}"
                return result
            
            file_content = response.content
            result['file_size_mb'] = round(len(file_content) / 1024 / 1024, 2)
            
            # Determine file type
            file_extension = Path(urlparse(excel_url).path).suffix.lower()
            
            content_parts = []
            
            if file_extension in ['.xlsx', '.xls'] and HAS_OPENPYXL:
                try:
                    # Process Excel file
                    from openpyxl import load_workbook
                    workbook = load_workbook(io.BytesIO(file_content), read_only=True)
                    
                    result['metadata']['sheet_names'] = workbook.sheetnames
                    result['metadata']['sheet_count'] = len(workbook.sheetnames)
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        sheet_data = []
                        
                        # Convert to list of lists
                        for row in sheet.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):  # Skip empty rows
                                sheet_data.append([str(cell) if cell is not None else '' for cell in row])
                        
                        if sheet_data:
                            # Convert to structured text
                            sheet_text = f"\n--- SHEET: {sheet_name} ---\n"
                            for row_num, row in enumerate(sheet_data):
                                if row_num == 0:  # Header row
                                    sheet_text += f"HEADERS: {' | '.join(row)}\n"
                                    sheet_text += "-" * 50 + "\n"
                                else:
                                    sheet_text += f"ROW {row_num}: {' | '.join(row)}\n"
                            
                            content_parts.append(sheet_text)
                            result['sheet_data'][sheet_name] = sheet_data
                    
                    workbook.close()
                    logger.info(f"Extracted {len(workbook.sheetnames)} sheets from Excel file")
                
                except Exception as e:
                    logger.warning(f"Excel processing failed: {e}")
            
            elif file_extension == '.csv' and HAS_PANDAS:
                try:
                    # Process CSV file
                    df = pd.read_csv(io.BytesIO(file_content))
                    
                    result['metadata']['row_count'] = len(df)
                    result['metadata']['column_count'] = len(df.columns)
                    result['metadata']['columns'] = list(df.columns)
                    
                    # Convert to structured text
                    csv_text = f"\n--- CSV DATA ---\n"
                    csv_text += f"COLUMNS: {' | '.join(df.columns)}\n"
                    csv_text += "-" * 50 + "\n"
                    
                    for index, row in df.iterrows():
                        row_text = f"ROW {index + 1}: {' | '.join([str(val) for val in row.values])}\n"
                        csv_text += row_text
                    
                    content_parts.append(csv_text)
                    result['sheet_data']['csv_data'] = df.to_dict('records')
                    
                    logger.info(f"Extracted CSV with {len(df)} rows and {len(df.columns)} columns")
                
                except Exception as e:
                    logger.warning(f"CSV processing failed: {e}")
            
            if content_parts:
                # Add metadata header
                metadata_header = f"""--- SPREADSHEET DOCUMENT METADATA ---
Source URL: {excel_url}
Extraction Date: {datetime.now().isoformat()}
File Size: {result['file_size_mb']} MB
File Type: {file_extension}
Sheets/Tables: {result['metadata'].get('sheet_count', 1)}
--- END METADATA ---

"""
                
                result['text_content'] = metadata_header + "\n\n".join(content_parts)
                result['success'] = True
                
                logger.info(f"Successfully extracted spreadsheet content: {len(result['text_content'])} characters")
            else:
                result['error'] = "No data could be extracted"
            
            return result
            
        except Exception as e:
            result['error'] = str(e)
            logger.error(f"Error extracting spreadsheet {excel_url}: {e}")
            return result

class EnhancedAUSTRACUpdatesScraper:
    """Enhanced AUSTRAC Updates scraper with LLM-optimized content extraction"""
    
    def __init__(self, run_type: str = "daily"):
        self.base_url = BASE_URL
        self.target_url = TARGET_URL
        self.data_folder = DATA_DIR
        self.data_folder.mkdir(exist_ok=True)
        
        # Determine run type
        self.run_type = run_type.lower()  # "daily" or "initial"
        self.max_pages = MAX_PAGES_DAILY if self.run_type == "daily" else MAX_PAGES_INITIAL
        
        self.json_file = JSON_PATH
        self.driver = None
        self.existing_hashes = set()
        self.doc_processor = DocumentProcessor()
        
        # Initialize tracking variables
        self._recent_articles_found = 0
        self._pages_without_recent = 0
        
        # Load existing data for deduplication
        self._load_existing_data()
        
        logger.info(f"Initialized scraper for {self.run_type.upper()} run (max {self.max_pages} pages)")
    
    def _load_existing_data(self):
        """Load existing data to prevent duplicates"""
        if self.json_file.exists():
            try:
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Handle both old format (list) and new format (dict with articles key)
                    if isinstance(data, list):
                        existing_data = data
                    else:
                        existing_data = data.get('articles', [])
                    
                    self.existing_hashes = {item.get('hash_id', '') for item in existing_data if item.get('hash_id')}
                    logger.info(f"Loaded {len(self.existing_hashes)} existing records")
            except Exception as e:
                logger.error(f"Error loading existing data: {e}")
                self.existing_hashes = set()
        else:
            logger.info("No existing data file found - starting fresh")
    
    def _generate_hash(self, url: str, headline: str, published_date: str) -> str:
        """Generate unique hash for article"""
        content = f"{url}_{headline}_{published_date}"
        return hashlib.sha256(content.encode('utf-8')).hexdigest()
    
    def _should_continue_scraping(self) -> bool:
        """Determine if scraping should continue based on run type and recent articles"""
        if self.run_type == "initial":
            return True  # Always continue for initial runs
        
        # For daily runs, check if we've found recent articles
        if self._recent_articles_found > 5:
            return True
        
        # Stop if we've gone through enough pages without recent content
        if self._pages_without_recent > 2:
            logger.info("Stopping daily run - no recent articles found in recent pages")
            return False
        
        return True
    
    def _is_recent_article(self, published_date: str) -> bool:
        """Check if article is recent (within last 7 days for daily runs)"""
        if self.run_type == "initial":
            return True
        
        try:
            # Parse the date
            parsed_date = self._parse_date(published_date)
            if not parsed_date:
                return True  # If we can't parse, assume it's recent
            
            # Check if within last 7 days
            cutoff_date = datetime.now() - timedelta(days=7)
            return parsed_date >= cutoff_date
        except:
            return True  # Default to including if unsure
    
    def _setup_driver(self):
        """Setup Chrome WebDriver with enhanced compatibility"""
        chrome_options = Options()
        
        # Essential stability options for Linux/WSL
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-software-rasterizer")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-features=TranslateUI")
        chrome_options.add_argument("--disable-ipc-flooding-protection")
        
        # Stealth options
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Set realistic window size
        chrome_options.add_argument("--window-size=1920,1080")
        
        # Updated user agent
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36")
        
        # Headless mode
        chrome_options.add_argument("--headless=new")
        
        # Performance optimizations
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-plugins")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_argument("--silent")
        
        try:
            # Try to find ChromeDriver
            possible_chromedriver_paths = [
                "/usr/bin/chromedriver",
                "/usr/local/bin/chromedriver", 
                "/snap/bin/chromedriver"
            ]
            
            chromedriver_path = None
            for path in possible_chromedriver_paths:
                if os.path.exists(path):
                    chromedriver_path = path
                    logger.info(f"Found ChromeDriver at: {path}")
                    break
            
            if chromedriver_path:
                service = Service(chromedriver_path)
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
                logger.info(f"Using ChromeDriver: {chromedriver_path}")
            else:
                self.driver = webdriver.Chrome(options=chrome_options)
                logger.info("Using ChromeDriver from PATH")
            
            # Set timeouts
            self.driver.implicitly_wait(10)
            self.driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
            
            # Execute script to remove webdriver property
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            logger.info("Chrome driver initialized successfully")
            return True
            
        except Exception as e:
            logger.error(f"Failed to initialize Chrome driver: {e}")
            return False
    
    def _is_driver_alive(self):
        """Check if driver is still responsive"""
        try:
            if self.driver is None:
                return False
            _ = self.driver.current_url
            return True
        except Exception:
            return False
    
    def _parse_date(self, date_str: str) -> Optional[datetime]:
        """Parse various date formats from AUSTRAC website"""
        try:
            # Clean the date string
            date_str = date_str.strip()
            
            # Try multiple date formats
            date_formats = [
                "%d %B %Y",      # 3 June 2025
                "%d %b %Y",      # 3 Jun 2025  
                "%B %d, %Y",     # June 3, 2025
                "%b %d, %Y",     # Jun 3, 2025
                "%Y-%m-%d",      # 2025-06-03
                "%d/%m/%Y",      # 03/06/2025
                "%m/%d/%Y"       # 06/03/2025
            ]
            
            for fmt in date_formats:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            
            logger.warning(f"Could not parse date: {date_str}")
            return None
            
        except Exception as e:
            logger.warning(f"Error parsing date '{date_str}': {e}")
            return None
    
    def _extract_structured_content_for_llm(self, soup: BeautifulSoup) -> str:
        """Extract and structure content optimally for LLM analysis"""
        content_parts = []
        
        # Extract headline
        headline = ""
        headline_selectors = ['.au-header-heading', 'h1', '.page-title', '.article-title', '.headline']
        for selector in headline_selectors:
            elem = soup.select_one(selector)
            if elem:
                headline = elem.get_text(strip=True)
                break
        
        if headline:
            content_parts.append(f"HEADLINE: {headline}")
        
        # Extract publication date
        date_info = ""
        date_selectors = ['time[datetime]', '.date', '.published-date', '.article-date']
        for selector in date_selectors:
            elem = soup.select_one(selector)
            if elem:
                date_info = elem.get_text(strip=True)
                datetime_attr = elem.get('datetime')
                if datetime_attr:
                    date_info += f" (ISO: {datetime_attr})"
                break
        
        if date_info:
            content_parts.append(f"PUBLICATION DATE: {date_info}")
        
        # Extract main content with structure preservation
        main_content = ""
        content_selectors = ['.body-copy', '.field--name-body', '.field__item', '.content', '.article-content', '.main-content', 'article']
        
        for selector in content_selectors:
            content_elem = soup.select_one(selector)
            if content_elem:
                # Remove unwanted elements
                for unwanted in content_elem.select('nav, .navigation, .breadcrumb, .share, .tags, .metadata, script, style, .visually-hidden'):
                    unwanted.decompose()
                
                # Process content with structure
                structured_content = []
                
                # Extract paragraphs
                paragraphs = content_elem.find_all('p')
                for i, p in enumerate(paragraphs, 1):
                    text = p.get_text(strip=True)
                    if text and len(text) > 10:  # Skip very short paragraphs
                        structured_content.append(f"PARAGRAPH {i}: {text}")
                
                # Extract lists
                lists = content_elem.find_all(['ul', 'ol'])
                for list_num, list_elem in enumerate(lists, 1):
                    list_items = list_elem.find_all('li')
                    if list_items:
                        structured_content.append(f"LIST {list_num}:")
                        for item_num, item in enumerate(list_items, 1):
                            item_text = item.get_text(strip=True)
                            if item_text:
                                structured_content.append(f"  - ITEM {item_num}: {item_text}")
                
                # Extract tables
                tables = content_elem.find_all('table')
                for table_num, table in enumerate(tables, 1):
                    structured_content.append(f"TABLE {table_num}:")
                    rows = table.find_all('tr')
                    for row_num, row in enumerate(rows, 1):
                        cells = row.find_all(['td', 'th'])
                        if cells:
                            cell_texts = [cell.get_text(strip=True) for cell in cells]
                            structured_content.append(f"  ROW {row_num}: {' | '.join(cell_texts)}")
                
                main_content = "\n".join(structured_content)
                break
        
        if main_content:
            content_parts.append(f"MAIN CONTENT:\n{main_content}")
        
        # Combine all parts
        final_content = "\n\n".join(content_parts)
        return DocumentProcessor.clean_text_for_llm(final_content)
    
    def _find_and_extract_documents(self, soup: BeautifulSoup, base_url: str) -> List[Dict[str, Any]]:
        """Find and extract content from linked PDFs and Excel/CSV files"""
        documents = []
        
        # Find all links
        links = soup.find_all('a', href=True)
        
        for link in links:
            href = link['href']
            
            # Make absolute URL
            if href.startswith('/'):
                full_url = urljoin(self.base_url, href)
            elif not href.startswith('http'):
                full_url = urljoin(base_url, href)
            else:
                full_url = href
            
            # Check file extension
            parsed_url = urlparse(full_url)
            file_path = parsed_url.path.lower()
            
            if file_path.endswith('.pdf'):
                logger.info(f"Found PDF: {full_url}")
                pdf_result = self.doc_processor.extract_pdf_content(full_url)
                if pdf_result['success']:
                    documents.append({
                        'type': 'PDF',
                        'url': full_url,
                        'link_text': link.get_text(strip=True),
                        'content': pdf_result['text_content'],
                        'metadata': pdf_result['metadata'],
                        'file_size_mb': pdf_result['file_size_mb'],
                        'page_count': pdf_result['page_count']
                    })
                else:
                    logger.warning(f"Failed to extract PDF {full_url}: {pdf_result.get('error')}")
            
            elif file_path.endswith(('.xlsx', '.xls', '.csv')):
                logger.info(f"Found spreadsheet: {full_url}")
                excel_result = self.doc_processor.extract_excel_content(full_url)
                if excel_result['success']:
                    documents.append({
                        'type': 'SPREADSHEET',
                        'url': full_url,
                        'link_text': link.get_text(strip=True),
                        'content': excel_result['text_content'],
                        'sheet_data': excel_result['sheet_data'],
                        'metadata': excel_result['metadata'],
                        'file_size_mb': excel_result['file_size_mb']
                    })
                else:
                    logger.warning(f"Failed to extract spreadsheet {full_url}: {excel_result.get('error')}")
        
        return documents
    
    def _find_contextual_links(self, soup: BeautifulSoup, base_url: str) -> List[Dict[str, Any]]:
        """Find links within textual content only (p, ul, li, etc.) - excludes navigation links"""
        contextual_links = []
        
        # Only look for links within content elements
        content_selectors = ['.body-copy', '.field--name-body', '.field__item', '.content', '.article-content', '.main-content', 'article']
        content_container = None
        
        for selector in content_selectors:
            content_container = soup.select_one(selector)
            if content_container:
                break
        
        if not content_container:
            # Fallback to look in textual elements if no main content found
            content_container = soup
        
        # Find links only within textual elements
        textual_elements = content_container.find_all(['p', 'ul', 'ol', 'li', 'div', 'span', 'blockquote', 'article'])
        
        # Exclude navigation and other non-content areas
        excluded_classes = [
            'nav', 'navigation', 'menu', 'breadcrumb', 'share', 'tags', 'metadata', 
            'header', 'footer', 'sidebar', 'widget', 'social', 'follow', 'subscribe',
            'contact', 'careers', 'login', 'search'
        ]
        
        excluded_texts = [
            'skip to', 'careers', 'contact us', 'home', 'business', 'subscribe', 
            'login', 'register', 'enrol', 'main content', 'austrac online',
            'new to austrac', 'your industry', 'banking', 'bookmakers'
        ]
        
        for element in textual_elements:
            # Skip elements that are likely navigation
            element_classes = ' '.join(element.get('class', [])).lower()
            if any(excluded in element_classes for excluded in excluded_classes):
                continue
            
            links = element.find_all('a', href=True)
            for link in links:
                href = link['href']
                link_text = link.get_text(strip=True)
                
                # Skip empty links or excluded text patterns
                if not link_text or any(excluded in link_text.lower() for excluded in excluded_texts):
                    continue
                
                # Skip anchor links and javascript
                if href.startswith('#') or href.startswith('javascript:') or href.startswith('mailto:'):
                    continue
                
                # Make absolute URL
                if href.startswith('/'):
                    href = urljoin(self.base_url, href)
                elif not href.startswith('http'):
                    href = urljoin(base_url, href)
                
                # Skip if same as current page
                if href == base_url:
                    continue
                
                # Only include if link text seems meaningful (not just navigation)
                if len(link_text) > 5 and not link_text.lower() in excluded_texts:
                    contextual_links.append({
                        'url': href,
                        'text': link_text,
                        'domain': urlparse(href).netloc
                    })
        
        # Remove duplicates while preserving order
        seen = set()
        unique_links = []
        for link in contextual_links:
            if link['url'] not in seen:
                seen.add(link['url'])
                unique_links.append(link)
        
        return unique_links[:10]  # Limit to first 10 meaningful links
    
    def _extract_articles_from_page(self) -> List[Dict]:
        """Extract articles from the current page using multiple selector strategies"""
        articles = []
        
        try:
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Try multiple selectors for article extraction
            article_elements = soup.select('.latest-news__card')
            
            if not article_elements:
                article_elements = soup.select('.views-row, .node, .view-content .item')
            
            if not article_elements:
                article_elements = soup.select('article, .news-item, .update-item, .content-item')
            
            if not article_elements:
                # Fallback: look for any elements with links and dates
                article_elements = soup.select('div:has(a):has(time), li:has(a):has(time)')
            
            if not article_elements:
                # Last resort: look for structured content in lists
                article_elements = soup.select('.view-content > div, .content-list > div, .updates-list > div')
            
            logger.info(f"Found {len(article_elements)} potential article elements")
            
            recent_count = 0
            for element in article_elements:
                try:
                    article_data = self._parse_article_element(element)
                    if article_data:
                        # Check if article is recent (for daily runs)
                        if self._is_recent_article(article_data.get('published_date', '')):
                            recent_count += 1
                        
                        articles.append(article_data)
                        logger.info(f"Extracted article: {article_data['headline'][:100]}...")
                        
                except Exception as e:
                    logger.warning(f"Error parsing article element: {e}")
                    continue
            
            # Track recent articles for daily run logic
            self._recent_articles_found += recent_count
            
            return articles
            
        except Exception as e:
            logger.error(f"Error extracting articles from page: {e}")
            return []
    
    def _parse_article_element(self, element) -> Optional[Dict]:
        """Parse individual article element with improved robustness"""
        try:
            # Extract headline and URL
            headline = None
            url = None
            
            # Enhanced headline selectors for AUSTRAC Updates
            headline_selectors = [
                '.latest-news__card-title a',
                '.latest-news__card-title',
                '.node-title a',
                '.views-field-title a',
                '.field--name-title a',
                'h1 a, h2 a, h3 a, h4 a, h5 a, h6 a',
                '.title a',
                '.headline a',
                'a[href*="/business/updates/"]',
                'a[href*="/news/"]',
                'a'
            ]
            
            for selector in headline_selectors:
                headline_elem = element.select_one(selector)
                if headline_elem:
                    headline = headline_elem.get_text(strip=True)
                    if headline_elem.name == 'a' and headline_elem.get('href'):
                        url = headline_elem.get('href')
                    elif headline_elem.find('a'):
                        url = headline_elem.find('a').get('href', '')
                    
                    if headline and len(headline) > 10:  # Ensure substantial headline
                        break
            
            if not headline:
                # Try without anchor tag
                for selector in ['.latest-news__card-title', '.node-title', '.views-field-title', '.field--name-title', 'h1, h2, h3, h4, h5, h6', '.title', '.headline']:
                    headline_elem = element.select_one(selector)
                    if headline_elem:
                        headline = headline_elem.get_text(strip=True)
                        if len(headline) > 10:
                            break
            
            if not headline or len(headline) < 10:
                return None
            
            # Look for URL if not found yet
            if not url:
                link_selectors = [
                    'a[href*="/business/updates/"]',
                    'a[href*="/news/"]',
                    'a[href*="/business/"]',
                    'a[href]'
                ]
                
                for selector in link_selectors:
                    link_elem = element.select_one(selector)
                    if link_elem:
                        url = link_elem.get('href')
                        break
            
            if not url:
                return None
            
            # Make URL absolute
            if url.startswith('/'):
                url = self.base_url + url
            elif not url.startswith('http'):
                url = self.base_url + '/' + url.lstrip('/')
            
            # Extract published date with enhanced selectors
            published_date = "Unknown"
            date_selectors = [
                'time',
                '.latest-news__card-date',
                '.views-field-created',
                '.field--name-created',
                '.date',
                '.published',
                '.field--name-field-article-dateline',
                '.post-date'
            ]
            
            for selector in date_selectors:
                date_elem = element.select_one(selector)
                if date_elem:
                    published_date = date_elem.get_text(strip=True)
                    # Also try datetime attribute
                    if not published_date and date_elem.get('datetime'):
                        published_date = date_elem.get('datetime')
                    if published_date and published_date != "Unknown":
                        break
            
            # Generate hash for deduplication
            hash_id = self._generate_hash(url, headline, published_date)
            
            # Check if already exists
            if hash_id in self.existing_hashes:
                logger.debug(f"Skipping existing article: {headline[:50]}...")
                return None
            
            return {
                'hash_id': hash_id,
                'headline': self._clean_text(headline),
                'url': url,
                'published_date': published_date,
                'scraped_date': datetime.now(timezone.utc).isoformat(),
                'category': self._extract_category(headline),
                'run_type': self.run_type
            }
            
        except Exception as e:
            logger.error(f"Error parsing element: {e}")
            return None
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize text content"""
        if not text:
            return ""
        
        # Remove excessive whitespace
        text = re.sub(r'\s+', ' ', text)
        # Remove special characters but keep essential punctuation
        text = re.sub(r'[^\w\s\.\,\;\:\!\?\-\(\)\[\]\{\}\"\'\/\\\@\#\$\%\&\*\+\=\<\>\~\`]', '', text)
        # Clean up spacing around punctuation
        text = re.sub(r'\s+([.!?,:;])', r'\1', text)
        text = re.sub(r'([.!?])\s*([A-Z])', r'\1 \2', text)
        
        return text.strip()
    
    def _extract_category(self, headline: str) -> str:
        """Extract category from headline"""
        headline_lower = headline.lower()
        
        categories = {
            'enforcement': ['penalty', 'infringement', 'compliance', 'audit', 'breach', 'violation', 'civil penalty', 'enforcement action'],
            'regulation': ['regulation', 'requirement', 'obligation', 'rule', 'standard', 'legislative', 'act amendment'],
            'guidance': ['guidance', 'update', 'information', 'clarification', 'advisory', 'notice'],
            'partnership': ['partnership', 'alliance', 'cooperation', 'joint', 'collaboration', 'mou'],
            'technology': ['crypto', 'digital', 'technology', 'fintech', 'blockchain', 'cryptocurrency', 'atm'],
            'industry': ['bank', 'casino', 'remitter', 'exchange', 'financial', 'gaming', 'betting'],
            'reform': ['reform', 'amendment', 'change', 'new law', 'legislative change'],
            'intelligence': ['intelligence', 'report', 'analysis', 'data', 'suspicious', 'typology'],
            'international': ['international', 'global', 'fatf', 'overseas', 'foreign', 'cross-border'],
            'education': ['forum', 'education', 'training', 'workshop', 'seminar', 'conference'],
            'scams': ['scam', 'fraud', 'illicit', 'criminal', 'money laundering'],
            'registration': ['registration', 'registered', 'provider', 'licensing']
        }
        
        for category, keywords in categories.items():
            if any(keyword in headline_lower for keyword in keywords):
                return category.title()
        
        return 'General'
    
    def _extract_article_content_enhanced(self, article: Dict) -> Dict:
        """Extract comprehensive content optimized for LLM analysis"""
        try:
            logger.info(f"Extracting enhanced content for: {article['headline'][:100]}...")
            
            # Check driver health
            if not self._is_driver_alive():
                logger.warning("Driver not responsive, reinitializing...")
                if not self._setup_driver():
                    return article
            
            self.driver.get(article['url'])
            time.sleep(random.uniform(2, 4))
            
            # Wait for page to load
            try:
                WebDriverWait(self.driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                # Additional wait for dynamic content
                time.sleep(2)
            except TimeoutException:
                logger.warning("Page load timeout")
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Extract structured content for LLM
            structured_content = self._extract_structured_content_for_llm(soup)
            
            # Extract linked documents
            linked_documents = self._find_and_extract_documents(soup, article['url'])
            
            # Extract contextual links from content only
            contextual_links = self._find_contextual_links(soup, article['url'])
            
            # Update article with enhanced content
            article['structured_content'] = structured_content
            article['linked_documents'] = linked_documents
            article['contextual_links'] = contextual_links
            article['document_count'] = len(linked_documents)
            article['total_content_length'] = len(structured_content) + sum(len(doc.get('content', '')) for doc in linked_documents)
            
            if structured_content:
                logger.info(f"Enhanced extraction complete: {article['total_content_length']} total characters, {len(linked_documents)} documents")
            else:
                logger.warning(f"No content extracted for: {article['headline'][:50]}...")
            
            return article
            
        except Exception as e:
            logger.error(f"Error in enhanced extraction for {article['url']}: {e}")
            return article

    def scrape_articles(self) -> List[Dict]:
        """Main scraping method with enhanced content extraction"""
        logger.info("="*80)
        logger.info(f"Starting AUSTRAC Updates enhanced scraping ({self.run_type.upper()} run)")
        logger.info("="*80)
        
        if not self._setup_driver():
            return []
        
        try:
            all_articles = []
            page_num = 0
            consecutive_empty_pages = 0
            max_empty_pages = 3 if self.run_type == "daily" else 5
            
            # Initialize tracking variables
            self._recent_articles_found = 0
            self._pages_without_recent = 0
            
            while (consecutive_empty_pages < max_empty_pages and 
                   page_num < self.max_pages and 
                   not shutdown_requested and
                   self._should_continue_scraping()):
                
                if page_num > 0:
                    url = f"{self.target_url}?page={page_num}"
                else:
                    url = self.target_url
                
                logger.info(f"Scraping page {page_num + 1}: {url}")
                
                # Check driver health
                if not self._is_driver_alive():
                    logger.warning("Driver not responsive, reinitializing...")
                    if not self._setup_driver():
                        break
                
                self.driver.get(url)
                
                # Wait for page to load
                try:
                    WebDriverWait(self.driver, 15).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                    time.sleep(random.uniform(2, 4))
                except TimeoutException:
                    logger.warning("Page load timeout")
                
                # Wait for articles to load
                try:
                    wait = WebDriverWait(self.driver, 15)
                    selectors_to_try = [
                        ".latest-news__card",
                        ".views-row",
                        ".news-item",
                        ".update-item",
                        "article"
                    ]
                    
                    articles_loaded = False
                    for selector in selectors_to_try:
                        try:
                            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                            logger.debug(f"Articles loaded with selector: {selector}")
                            articles_loaded = True
                            break
                        except TimeoutException:
                            continue
                    
                    if not articles_loaded:
                        logger.warning("No articles found with any selector")
                
                except TimeoutException:
                    logger.warning("Timeout waiting for articles to load")
                
                page_articles = self._extract_articles_from_page()
                
                if not page_articles:
                    consecutive_empty_pages += 1
                    logger.info(f"Empty page ({consecutive_empty_pages}/{max_empty_pages})")
                else:
                    consecutive_empty_pages = 0
                    all_articles.extend(page_articles)
                    logger.info(f"Found {len(page_articles)} articles on page {page_num + 1}")
                
                # For daily runs, check if we found recent articles
                if self.run_type == "daily":
                    recent_articles = [a for a in page_articles if self._is_recent_article(a.get('published_date', ''))]
                    if not recent_articles:
                        self._pages_without_recent += 1
                    else:
                        self._pages_without_recent = 0
                
                # Check if there are more pages
                soup = BeautifulSoup(self.driver.page_source, 'html.parser')
                pagination_selectors = [
                    '.pager__item--next',
                    '.pagination .next',
                    'a[rel="next"]',
                    '.next-page',
                    '.pager .pager-next'
                ]
                
                has_next_page = False
                for selector in pagination_selectors:
                    next_link = soup.select_one(selector)
                    if next_link and not next_link.get('disabled'):
                        has_next_page = True
                        break
                
                if not has_next_page:
                    logger.info("No more pages found")
                    break
                
                page_num += 1
                time.sleep(random.uniform(2, 4))
            
            if not all_articles:
                logger.info("No new articles found")
                return []
            
            logger.info(f"Found {len(all_articles)} articles. Starting enhanced content extraction...")
            
            # Enhanced content extraction for each article
            enriched_articles = []
            for i, article in enumerate(all_articles):
                if shutdown_requested:
                    break
                
                logger.info(f"Processing article {i+1}/{len(all_articles)}: {article['headline'][:80]}...")
                
                try:
                    enriched_article = self._extract_article_content_enhanced(article)
                    enriched_articles.append(enriched_article)
                    
                    # Log extraction results
                    total_chars = enriched_article.get('total_content_length', 0)
                    doc_count = enriched_article.get('document_count', 0)
                    logger.info(f"  Content: {total_chars:,} chars, Documents: {doc_count}")
                    
                except Exception as e:
                    logger.error(f"Error processing article {i+1}: {e}")
                    enriched_articles.append(article)  # Add original if enhancement fails
                
                time.sleep(random.uniform(2, 4))
            
            logger.info("="*80)
            logger.info(f"Enhanced extraction complete: {len(enriched_articles)} articles processed")
            
            # Calculate summary
            total_documents = sum(article.get('document_count', 0) for article in enriched_articles)
            total_content = sum(article.get('total_content_length', 0) for article in enriched_articles)
            
            logger.info(f"Total linked documents processed: {total_documents}")
            logger.info(f"Total content extracted: {total_content:,} characters")
            logger.info("="*80)
            
            return enriched_articles
            
        finally:
            if self.driver:
                self.driver.quit()
                logger.info("Browser closed")

    def save_articles(self, new_articles: List[Dict]):
        """Save articles to JSON only (optimized for LLM analysis)"""
        if not new_articles:
            logger.info("No new articles to save")
            return
        
        # Load existing data
        existing_articles = []
        if self.json_file.exists():
            try:
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Handle both old format (list) and new format (dict with articles key)
                    if isinstance(data, list):
                        existing_articles = data
                    else:
                        existing_articles = data.get('articles', [])
            except Exception as e:
                logger.error(f"Error loading existing data: {e}")
        
        # Filter truly new articles
        existing_hash_ids = {article.get('hash_id', '') for article in existing_articles}
        actually_new = [article for article in new_articles 
                       if article.get('hash_id', '') not in existing_hash_ids]
        
        if not actually_new:
            logger.info("No genuinely new articles - all were duplicates")
            return
        
        # Merge and sort
        all_articles = existing_articles + actually_new
        try:
            all_articles.sort(key=lambda x: x.get('scraped_date', ''), reverse=True)
        except Exception as e:
            logger.warning(f"Could not sort by date: {e}")
        
        # Save enhanced JSON structure
        output_data = {
            'metadata': {
                'last_updated': datetime.now(timezone.utc).isoformat(),
                'run_type': self.run_type,
                'total_articles': len(all_articles),
                'new_articles_added': len(actually_new),
                'extraction_capabilities': {
                    'pdf_extraction': HAS_PYMUPDF or HAS_PDFPLUMBER or HAS_PYPDF2,
                    'excel_extraction': HAS_OPENPYXL and HAS_PANDAS,
                    'structured_content': True,
                    'llm_optimized': True
                }
            },
            'articles': all_articles
        }
        
        try:
            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Saved {len(all_articles)} total articles to {self.json_file}")
            logger.info(f"Added {len(actually_new)} new articles")
            
            # Log summary of new additions
            if actually_new:
                logger.info("NEW ARTICLES ADDED:")
                for article in actually_new:
                    doc_count = article.get('document_count', 0)
                    content_len = article.get('total_content_length', 0)
                    logger.info(f"  • {article.get('category', 'General')}: {article.get('headline', 'No title')[:80]}...")
                    if doc_count > 0:
                        logger.info(f"    Documents: {doc_count}, Content: {content_len:,} chars")
            
        except Exception as e:
            logger.error(f"Error saving JSON: {e}")

def main():
    """Main execution function with run type detection"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Enhanced AUSTRAC Updates Scraper')
    parser.add_argument('--run-type', choices=['daily', 'initial'], default='daily',
                        help='Type of run: daily (recent articles) or initial (comprehensive)')
    parser.add_argument('--max-pages', type=int, help='Override maximum pages to scrape')
    
    args = parser.parse_args()
    
    try:
        # Determine run type
        run_type = args.run_type
        
        # Auto-detect initial run if no existing data
        if not JSON_PATH.exists() and run_type == 'daily':
            logger.info("No existing data found, switching to initial run")
            run_type = 'initial'
        
        scraper = EnhancedAUSTRACUpdatesScraper(run_type=run_type)
        
        # Override max pages if specified
        if args.max_pages:
            scraper.max_pages = args.max_pages
            logger.info(f"Override: Max pages set to {args.max_pages}")
        
        logger.info("="*80)
        logger.info(f"AUSTRAC UPDATES ENHANCED SCRAPER STARTED ({run_type.upper()} RUN)")
        logger.info("Features: LLM-optimized content, PDF/Excel extraction, structured output")
        logger.info("="*80)
        
        new_articles = scraper.scrape_articles()
        
        if new_articles:
            scraper.save_articles(new_articles)
            
            print("="*60)
            print("AUSTRAC UPDATES ENHANCED SCRAPING SUMMARY")
            print("="*60)
            print(f"Run type: {run_type.upper()}")
            print(f"Articles processed: {len(new_articles)}")
            total_docs = sum(article.get('document_count', 0) for article in new_articles)
            if total_docs > 0:
                print(f"Documents extracted: {total_docs}")
            print("="*60)
            print(f"Output: {JSON_PATH} (LLM-optimized JSON)")
            print("="*60)
            
        else:
            print("No new articles found - database is up to date")
            
    except KeyboardInterrupt:
        logger.info("Scraping interrupted by user")
    except Exception as e:
        logger.error(f"Scraping failed: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise

if __name__ == "__main__":
    main()
            