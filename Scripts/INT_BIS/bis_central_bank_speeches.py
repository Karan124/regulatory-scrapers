#!/usr/bin/env python3
"""
BIS Central Bank Speeches Scraper
Comprehensive scraper for Bank for International Settlements Central Bank Speeches
Extracts speeches, PDFs, Excel files, and metadata with intelligent content deduplication

Requirements:
pip install requests beautifulsoup4 pandas fake-useragent PyMuPDF openpyxl tiktoken lxml selenium webdriver-manager
"""

import requests
import json
import os
import time
import re
import zipfile
import io
import string
from datetime import datetime
from urllib.parse import urljoin, urlparse
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from difflib import SequenceMatcher

# Third-party imports
from bs4 import BeautifulSoup
import pandas as pd
from fake_useragent import UserAgent
import fitz  # PyMuPDF for PDF extraction
import openpyxl
from openpyxl import load_workbook
import tiktoken

# Selenium imports for JavaScript handling
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("Selenium not available. Install with: pip install selenium webdriver-manager")

# Configuration
BASE_URL = "https://www.bis.org"
START_URL = "https://www.bis.org/cbspeeches/index.htm"
DATA_DIR = Path("data")
OUTPUT_FILE = DATA_DIR / "bis_speeches.json"

# Set MAX_PAGE for different run types
MAX_PAGES = 3  # Change to None for full scrape, or set to desired number

class BISSpeechesScraper:
    def __init__(self, use_selenium: bool = True):
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.session = requests.Session()
        self.ua = UserAgent()
        self.driver = None
        self.existing_urls = set()
        self.encoding = tiktoken.get_encoding("cl100k_base")  # GPT-4 tokenizer
        
        # Create data directory
        DATA_DIR.mkdir(exist_ok=True)
        
        # Initialize components
        self.setup_session()
        self.load_existing_data()
        
        if self.use_selenium:
            self.setup_selenium()
    
    def setup_session(self):
        """Setup session with realistic headers and cookies"""
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0'
        })
        
        # Visit root page first to collect cookies
        try:
            response = self.session.get(BASE_URL, timeout=10)
            print(f"Root page status: {response.status_code}")
            time.sleep(2)
        except Exception as e:
            print(f"Warning: Could not visit root page: {e}")
    
    def setup_selenium(self):
        """Setup Selenium WebDriver for JavaScript handling"""
        if not SELENIUM_AVAILABLE:
            return
        
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument(f'--user-agent={self.ua.random}')
        
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            print("Selenium WebDriver initialized successfully")
        except Exception as e:
            print(f"Failed to initialize Selenium: {e}")
            self.use_selenium = False
    
    def get_page_content_selenium(self, url: str) -> Optional[str]:
        """Get page content using Selenium for JavaScript sites"""
        if not self.driver:
            return None
        
        try:
            self.driver.get(url)
            
            # Wait for content to load - flexible waiting
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.any_of(
                        EC.presence_of_element_located((By.CLASS_NAME, "documentList")),
                        EC.presence_of_element_located((By.CLASS_NAME, "item")),
                        EC.presence_of_element_located((By.TAG_NAME, "table"))
                    )
                )
            except:
                pass  # Continue anyway if elements not found
            
            time.sleep(3)  # Additional wait for dynamic content
            return self.driver.page_source
            
        except Exception as e:
            print(f"Selenium error for {url}: {e}")
            return None
    
    def load_existing_data(self):
        """Load existing scraped data to avoid duplicates"""
        if OUTPUT_FILE.exists():
            try:
                with open(OUTPUT_FILE, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                    self.existing_urls = {item.get('url', '') for item in existing_data if item.get('url')}
                print(f"Loaded {len(self.existing_urls)} existing speeches")
            except Exception as e:
                print(f"Warning: Could not load existing data: {e}")
                self.existing_urls = set()
        else:
            print("No existing data file found")
    
    def count_tokens(self, text: str) -> int:
        """Count tokens using tiktoken encoding"""
        if not text:
            return 0
        try:
            return len(self.encoding.encode(str(text)))
        except Exception:
            # Fallback to word count approximation
            return int(len(str(text).split()) * 1.3)
    
    def is_text_corrupted(self, text: str) -> bool:
        """Definitive corruption detection for text content"""
        if not text or len(text) < 50:
            return True
        
        # Count different character types
        letters = sum(1 for c in text if c.isalpha())
        digits = sum(1 for c in text if c.isdigit())
        spaces = sum(1 for c in text if c.isspace())
        punctuation = sum(1 for c in text if c in '.,!?;:()[]{}"\'-')
        total_chars = len(text)
        
        # Calculate ratios
        letter_ratio = letters / total_chars
        normal_chars = letters + digits + spaces + punctuation
        normal_ratio = normal_chars / total_chars
        
        # Text is corrupted if:
        # 1. Less than 40% letters
        # 2. Less than 70% normal characters
        # 3. Contains corruption patterns
        if letter_ratio < 0.4 or normal_ratio < 0.7:
            return True
        
        # Check for specific corruption patterns
        corruption_indicators = [
            len(re.findall(r'[^a-zA-Z0-9\s.,!?;:()\[\]{}"\'\\-]{3,}', text)) > 5,  # Strange char sequences
            len(re.findall(r'[A-Z]{15,}', text)) > 2,  # Long caps sequences
            len(re.findall(r'[0-9]{10,}', text)) > 2,  # Long number sequences
            len(re.findall(r'[^\x20-\x7E]{5,}', text)) > 0,  # Non-printable ASCII sequences
        ]
        
        return any(corruption_indicators)
    
    def safe_request(self, url: str, retries: int = 3) -> Optional[requests.Response]:
        """Make a safe request with retries and error handling"""
        for attempt in range(retries):
            try:
                # Update headers for each request
                self.session.headers.update({
                    'Referer': BASE_URL,
                    'Sec-Fetch-Site': 'same-origin' if BASE_URL in url else 'cross-site'
                })
                
                response = self.session.get(url, timeout=15)
                if response.status_code == 200:
                    return response
                elif response.status_code == 429:
                    wait_time = (attempt + 1) * 5
                    print(f"Rate limited. Waiting {wait_time}s...")
                    time.sleep(wait_time)
                else:
                    print(f"HTTP {response.status_code} for {url}")
                    
            except Exception as e:
                print(f"Request error (attempt {attempt + 1}): {e}")
                if attempt < retries - 1:
                    time.sleep((attempt + 1) * 2)
        
        return None
    
    def extract_pdf_text(self, pdf_url: str) -> str:
        """Extract text from PDF files with definitive corruption handling"""
        try:
            response = self.safe_request(pdf_url)
            if not response:
                return ""
            
            print(f"  Attempting PDF extraction: {pdf_url}")
            pdf_document = fitz.open(stream=response.content, filetype="pdf")
            
            # Try to extract text from all pages
            all_text_methods = []
            
            for page_num in range(min(len(pdf_document), 10)):  # Limit to first 10 pages for efficiency
                page = pdf_document.load_page(page_num)
                
                # Method 1: Standard extraction
                try:
                    text1 = page.get_text()
                    if text1 and not self.is_text_corrupted(text1):
                        all_text_methods.append(text1)
                except:
                    pass
                
                # Method 2: Text blocks
                try:
                    blocks = page.get_text("blocks")
                    text2 = " ".join([block[4] for block in blocks if len(block) > 4])
                    if text2 and not self.is_text_corrupted(text2):
                        all_text_methods.append(text2)
                except:
                    pass
            
            pdf_document.close()
            
            # Choose best extraction method
            if all_text_methods:
                # Use the longest non-corrupted text
                best_text = max(all_text_methods, key=len)
                
                # Final corruption check
                if self.is_text_corrupted(best_text):
                    print(f"  PDF text failed corruption check, rejecting")
                    return ""
                
                print(f"  Successfully extracted {len(best_text)} characters from PDF")
                return best_text
            else:
                print(f"  No clean text could be extracted from PDF")
                return ""
                
        except Exception as e:
            print(f"PDF extraction error for {pdf_url}: {e}")
            return ""
    
    def extract_excel_data(self, excel_url: str) -> str:
        """Extract data from Excel files"""
        try:
            response = self.safe_request(excel_url)
            if not response:
                return ""
            
            # Try pandas first
            try:
                excel_file = pd.ExcelFile(io.BytesIO(response.content))
                all_data = []
                
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    if not df.empty:
                        all_data.append(f"Sheet: {sheet_name}\n{df.to_string(index=False)}")
                
                return "\n\n".join(all_data)
                
            except Exception:
                # Fallback to openpyxl
                workbook = load_workbook(io.BytesIO(response.content), data_only=True)
                all_data = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_data = []
                    
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            row_data = [str(cell) if cell is not None else '' for cell in row]
                            sheet_data.append('\t'.join(row_data))
                    
                    if sheet_data:
                        all_data.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_data))
                
                return "\n\n".join(all_data)
                
        except Exception as e:
            print(f"Excel extraction error for {excel_url}: {e}")
            return ""
    
    def extract_zip_contents(self, zip_url: str) -> str:
        """Extract text content from ZIP files"""
        try:
            response = self.safe_request(zip_url)
            if not response:
                return ""
            
            with zipfile.ZipFile(io.BytesIO(response.content)) as zip_file:
                extracted_content = []
                
                for file_info in zip_file.infolist():
                    if file_info.filename.lower().endswith(('.txt', '.csv', '.json', '.md')):
                        try:
                            content = zip_file.read(file_info).decode('utf-8', errors='ignore')
                            extracted_content.append(f"File: {file_info.filename}\n{content}")
                        except Exception as e:
                            print(f"Error extracting {file_info.filename}: {e}")
                
                return "\n\n".join(extracted_content)
                
        except Exception as e:
            print(f"ZIP extraction error for {zip_url}: {e}")
            return ""
    
    def extract_tables(self, soup: BeautifulSoup) -> str:
        """Extract table data from soup"""
        try:
            tables = soup.find_all('table')
            table_data = []
            
            for i, table in enumerate(tables):
                rows = []
                for row in table.find_all('tr'):
                    cells = []
                    for cell in row.find_all(['td', 'th']):
                        cell_text = cell.get_text(strip=True)
                        cells.append(cell_text)
                    if cells and any(cell.strip() for cell in cells):  # Skip empty rows
                        rows.append('\t'.join(cells))
                
                if rows:
                    table_data.append(f"Table {i+1}:\n" + '\n'.join(rows))
            
            return "\n\n".join(table_data)
            
        except Exception as e:
            print(f"Table extraction error: {e}")
            return ""
    
    def extract_related_links(self, soup: BeautifulSoup, base_url: str) -> List[str]:
        """Extract related links from speech content"""
        try:
            content_div = soup.find('div', id='cmsContent')
            if not content_div:
                content_div = soup.find('div', {'id': 'center', 'role': 'main'})
            if not content_div:
                content_div = soup
            
            links = []
            for link in content_div.find_all('a', href=True):
                href = link['href']
                
                # Skip irrelevant links
                skip_patterns = ['mailto:', 'javascript:', '#', 'about', 'contact', 'terms', 'privacy', 'subscribe']
                if any(skip in href.lower() for skip in skip_patterns):
                    continue
                
                # Convert relative URLs to absolute
                full_url = urljoin(base_url, href)
                
                # Only include relevant domains
                relevant_domains = ['bis.org', 'fsb.org', 'iosco.org', 'bcbs.org', 'imf.org', 'worldbank.org']
                if any(domain in full_url for domain in relevant_domains):
                    links.append(full_url)
            
            return list(set(links))  # Remove duplicates
            
        except Exception as e:
            print(f"Link extraction error: {e}")
            return []
    
    def text_similarity(self, text1: str, text2: str) -> float:
        """Calculate similarity between two texts using SequenceMatcher"""
        if not text1 or not text2:
            return 0.0
        
        # Basic normalization
        text1_norm = re.sub(r'\s+', ' ', text1.lower().strip())
        text2_norm = re.sub(r'\s+', ' ', text2.lower().strip())
        
        if not text1_norm or not text2_norm:
            return 0.0
        
        return SequenceMatcher(None, text1_norm, text2_norm).ratio()
    
    def choose_canonical_content(self, web_text: str, pdf_text: str) -> Tuple[str, str]:
        """
        Choose the canonical content between web and PDF text
        Returns: (canonical_content, source_info)
        """
        # Clean and validate inputs
        web_clean = web_text.strip() if web_text else ""
        pdf_clean = pdf_text.strip() if pdf_text else ""
        
        # Check for corruption
        web_corrupted = self.is_text_corrupted(web_clean) if web_clean else True
        pdf_corrupted = self.is_text_corrupted(pdf_clean) if pdf_clean else True
        
        # Decision logic prioritizing clean content
        if web_corrupted and pdf_corrupted:
            return "", "No clean content available"
        
        if web_corrupted and not pdf_corrupted:
            return pdf_clean, "PDF only (web corrupted)"
        
        if not web_corrupted and pdf_corrupted:
            return web_clean, "Web only (PDF corrupted)"
        
        # Both are clean - use similarity analysis
        web_len = len(web_clean)
        pdf_len = len(pdf_clean)
        similarity = self.text_similarity(web_clean, pdf_clean)
        
        print(f"  Content comparison: Web={web_len} chars, PDF={pdf_len} chars, Similarity={similarity:.2f}")
        
        # High similarity - choose longer
        if similarity > 0.8:
            if pdf_len > web_len:
                return pdf_clean, f"PDF preferred (similar content, PDF longer: {pdf_len} vs {web_len})"
            else:
                return web_clean, f"Web preferred (similar content, Web longer/equal: {web_len} vs {pdf_len})"
        
        # Moderate similarity with significant length difference - likely truncation
        elif similarity > 0.5 and pdf_len > web_len * 1.5:
            return pdf_clean, f"PDF preferred (Web likely truncated: {web_len} -> {pdf_len})"
        
        # Low similarity or unclear - choose longer content
        elif pdf_len > web_len * 1.2:
            return pdf_clean, f"PDF preferred (different content, PDF longer: {pdf_len} vs {web_len})"
        else:
            return web_clean, f"Web preferred (different content, Web longer/similar: {web_len} vs {pdf_len})"
    
    def parse_speech_from_index(self, speech_element) -> Optional[Dict[str, Any]]:
        """Parse speech metadata from index page element"""
        try:
            # Extract date - multiple strategies
            date_elem = speech_element.find('td', class_='item_date')
            if not date_elem:
                date_elem = speech_element.find('td')
            
            published_date = date_elem.get_text(strip=True) if date_elem else ""
            
            # Extract title and URL
            title_link = speech_element.find('a')
            if not title_link:
                return None
            
            headline = title_link.get_text(strip=True)
            url = urljoin(BASE_URL, title_link['href'])
            
            # Extract speaker from various possible locations
            speaker = ""
            speech_info = speech_element.get_text()
            
            # Common patterns for speaker extraction
            speaker_patterns = [
                r'by\s+([^,\n]+?)(?:,|\n|$)',
                r'Speaker:\s*([^,\n]+?)(?:,|\n|$)',
                r'([^,\n]+?),\s*(?:Governor|Deputy Governor|Director|President|Chairman|Chair)',
            ]
            
            for pattern in speaker_patterns:
                match = re.search(pattern, speech_info, re.IGNORECASE)
                if match:
                    speaker = match.group(1).strip()
                    break
            
            # Extract category/institution
            category = ""
            category_elem = speech_element.find('span')
            if category_elem:
                category = category_elem.get_text(strip=True)
            else:
                # Try to extract from text
                if 'central bank' in speech_info.lower():
                    category = "Central Bank Speech"
                elif 'bis' in speech_info.lower():
                    category = "BIS Speech"
                else:
                    category = "Speech"
            
            return {
                'headline': headline,
                'url': url,
                'published_date': published_date,
                'speaker': speaker,
                'category': category
            }
            
        except Exception as e:
            print(f"Error parsing speech element: {e}")
            return None
    
    def scrape_speech_content(self, speech_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Scrape full content from speech page with intelligent PDF handling"""
        url = speech_data['url']
        
        # Skip if already scraped
        if url in self.existing_urls:
            print(f"Skipping already scraped: {speech_data['headline']}")
            return None
        
        print(f"Scraping: {speech_data['headline']}")
        
        # Check if this is a direct PDF link
        if url.lower().endswith('.pdf'):
            print(f"  Direct PDF link detected: {url}")
            
            # Extract PDF content directly
            pdf_text = self.extract_pdf_text(url)
            
            # Calculate token count
            token_count = self.count_tokens(pdf_text)
            
            # Build final speech data for PDF-only content
            return {
                **speech_data,
                'scraped_date': datetime.now().isoformat(),
                'content': pdf_text if pdf_text else f"Speech Document: {speech_data['headline']}\n\nPDF content could not be extracted cleanly.",
                'token_count': token_count,
                'related_links': [],
                'associated_image': "",
                'attachments': {
                    'pdf_text': pdf_text,
                    'excel_data': '',
                    'tables': '',
                    'charts': ''
                }
            }
        
        # Regular web page processing
        response = self.safe_request(url)
        if not response:
            return None
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extract main content
        content_div = soup.find('div', {'id': 'center', 'role': 'main'})
        if not content_div:
            content_div = soup.find('div', id='cmsContent')
        if not content_div:
            content_div = soup.find('div', class_='content')
        
        web_content = ""
        if content_div:
            # Remove navigation and irrelevant sections
            for unwanted in content_div.find_all(['nav', 'aside', 'header', 'footer']):
                unwanted.decompose()
            
            # Remove child documents section
            child_docs = content_div.find('div', class_='childdocsgroup')
            if child_docs:
                child_docs.decompose()
            
            # Extract text
            web_content = content_div.get_text(separator='\n', strip=True)
        
        # Extract speaker if not found in index
        if not speech_data.get('speaker'):
            speaker_patterns = [
                r'by\s+([^,\n]+?)(?:,|\n)',
                r'Speaker:\s*([^,\n]+?)(?:,|\n)',
                r'([^,\n]+?),\s*(?:Governor|Deputy Governor|Director|President|Chairman|Chair)',
            ]
            
            for pattern in speaker_patterns:
                match = re.search(pattern, web_content, re.IGNORECASE)
                if match:
                    speech_data['speaker'] = match.group(1).strip()
                    break
        
        # Extract associated image
        img_elem = soup.find('img', src=True)
        associated_image = ""
        if img_elem and 'src' in img_elem.attrs:
            img_src = img_elem['src']
            if not img_src.startswith('data:'):
                associated_image = urljoin(BASE_URL, img_src)
        
        # Extract tables
        tables_text = self.extract_tables(soup)
        
        # Extract related links
        related_links = self.extract_related_links(soup, url)
        
        # Process attachments and find PDFs
        attachments = {
            'pdf_text': '',
            'excel_data': '',
            'tables': tables_text,
            'charts': ''
        }
        
        pdf_text = ""
        pdf_urls = []
        
        # Find all file links
        file_links = soup.find_all('a', href=True)
        for link in file_links:
            href = link.get('href', '').strip()
            if not href:
                continue
            
            file_url = urljoin(BASE_URL, href)
            file_ext = href.lower()
            
            if file_ext.endswith('.pdf'):
                print(f"  Found PDF: {file_url}")
                pdf_urls.append(file_url)
                pdf_content = self.extract_pdf_text(file_url)
                if pdf_content:  # Only add if we got clean content
                    attachments['pdf_text'] += pdf_content + "\n\n"
                    if not pdf_text:  # Use first valid PDF for comparison
                        pdf_text = pdf_content
                time.sleep(1)  # Be respectful
                
            elif file_ext.endswith(('.xlsx', '.xls')):
                print(f"  Extracting Excel: {file_url}")
                attachments['excel_data'] += self.extract_excel_data(file_url) + "\n\n"
                time.sleep(1)
                
            elif file_ext.endswith('.zip'):
                print(f"  Extracting ZIP: {file_url}")
                attachments['excel_data'] += self.extract_zip_contents(file_url) + "\n\n"
                time.sleep(1)
        
        # Choose canonical content using intelligent logic
        canonical_content, content_source = self.choose_canonical_content(web_content, pdf_text)
        print(f"  Content decision: {content_source}")
        
        # Calculate total token count across chosen content and attachments
        all_text = canonical_content + attachments['pdf_text'] + attachments['excel_data'] + attachments['tables']
        token_count = self.count_tokens(all_text)
        
        # Build final speech data
        return {
            **speech_data,
            'scraped_date': datetime.now().isoformat(),
            'content': canonical_content,
            'token_count': token_count,
            'related_links': related_links,
            'associated_image': associated_image,
            'attachments': attachments
        }
    
    def get_speeches_from_page(self, url: str) -> List[Dict[str, Any]]:
        """Extract all speeches from an index page"""
        # Try Selenium first if available
        html_content = None
        if self.use_selenium:
            print("Using Selenium to get page content...")
            html_content = self.get_page_content_selenium(url)
        
        # Fallback to requests if Selenium fails or not available
        if not html_content:
            print("Using requests to get page content...")
            response = self.safe_request(url)
            if not response:
                return []
            html_content = response.text
        
        soup = BeautifulSoup(html_content, 'html.parser')
        speeches = []
        
        # Multiple strategies to find speech rows
        speech_rows = []
        
        # Strategy 1: Look for item classes (like BIS news)
        speech_rows = soup.find_all('tr', class_='item even') + soup.find_all('tr', class_='item odd')
        
        # Strategy 2: Look in main table structures
        if not speech_rows:
            tables = soup.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                # Filter for rows that contain links
                for row in rows:
                    if row.find('a', href=True):
                        speech_rows.append(row)
        
        # Strategy 3: Look for any row with date pattern
        if not speech_rows:
            date_pattern = re.compile(r'\d{1,2}\s+\w{3}\s+\d{4}')  # e.g., "15 Sep 2025"
            for tr in soup.find_all('tr'):
                if tr.find('td') and date_pattern.search(tr.get_text()) and tr.find('a'):
                    speech_rows.append(tr)
        
        print(f"Found {len(speech_rows)} speech rows in HTML")
        
        for row in speech_rows:
            speech_data = self.parse_speech_from_index(row)
            if speech_data:
                speeches.append(speech_data)
        
        return speeches
    
    def get_page_info(self, soup: BeautifulSoup) -> Tuple[int, int]:
        """Extract current page number and total pages from pagination"""
        try:
            # Look for pagination info in format "Page X of Y"
            page_info = soup.find('div', class_='pageof')
            if page_info:
                page_text = page_info.get_text()
                match = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', page_text)
                if match:
                    current_page = int(match.group(1))
                    total_pages = int(match.group(2))
        except Exception as e:
            print(f"Error extracting page info: {e}")
        
        return 1, 1  # Default fallback
    
    def construct_page_url(self, page_number: int) -> str:
        """Construct URL for a specific page number"""
        base_url = START_URL.split('?')[0]
        if page_number == 1:
            return base_url
        else:
            return f"{base_url}?cbspeeches_page={page_number}"
    
    def scrape_all_pages(self) -> List[Dict[str, Any]]:
        """Scrape speeches from all pages with simple, working pagination"""
        all_speeches = []
        page_number = 1
        
        while MAX_PAGES is None or page_number <= MAX_PAGES:
            # Construct URL for current page - EXACT pattern from your example
            if page_number == 1:
                current_url = "https://www.bis.org/cbspeeches/index.htm"
            else:
                current_url = f"https://www.bis.org/cbspeeches/index.htm?cbspeeches_page={page_number}"
            
            print(f"\n=== Scraping Page {page_number} ===")
            print(f"URL: {current_url}")
            
            # Get speeches from current page
            speeches = self.get_speeches_from_page(current_url)
            print(f"Found {len(speeches)} speeches on page {page_number}")
            
            # If no speeches found, we've likely reached the end or there's an issue
            if not speeches:
                print("No speeches found on this page")
                # Check if this is page 1 (which should always have content)
                if page_number == 1:
                    print("ERROR: No speeches found on page 1 - check scraping logic")
                    break
                else:
                    print("Reached end of available pages")
                    break
            
            # Scrape full content for each speech
            speeches_processed = 0
            for i, speech in enumerate(speeches, 1):
                print(f"  Processing speech {i}/{len(speeches)}: {speech['headline'][:60]}...")
                full_speech = self.scrape_speech_content(speech)
                if full_speech:
                    all_speeches.append(full_speech)
                    speeches_processed += 1
                
                time.sleep(2)  # Be respectful to the server
            
            print(f"Successfully processed {speeches_processed}/{len(speeches)} speeches from page {page_number}")
            
            # Check if we should continue to next page
            if MAX_PAGES is not None and page_number >= MAX_PAGES:
                print(f"Reached MAX_PAGES limit ({MAX_PAGES})")
                break
            
            # Also check if we got fewer speeches than expected (might indicate last page)
            # But continue anyway as page size might vary
            
            # Move to next page
            page_number += 1
            print(f"Moving to page {page_number}")
            time.sleep(3)  # Pause between pages
        
        return all_speeches
    
    def save_data(self, speeches: List[Dict[str, Any]]):
        """Save scraped data to JSON file with deduplication"""
        # Load existing data
        existing_data = []
        if OUTPUT_FILE.exists():
            try:
                with open(OUTPUT_FILE, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except Exception as e:
                print(f"Error loading existing data: {e}")
        
        # Merge new speeches (avoid duplicates by URL)
        existing_urls = {item.get('url') for item in existing_data if item.get('url')}
        new_speeches = [speech for speech in speeches if speech.get('url') not in existing_urls]
        
        combined_data = existing_data + new_speeches
        
        # Sort by published date (newest first)
        try:
            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, '%d %b %Y')
                except:
                    return datetime.min
            
            combined_data.sort(key=lambda x: parse_date(x.get('published_date', '')), reverse=True)
        except Exception as e:
            print(f"Could not sort by date: {e}")
        
        # Save to file
        with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
            json.dump(combined_data, f, indent=2, ensure_ascii=False)
        
        print(f"\nSaved {len(new_speeches)} new speeches ({len(combined_data)} total)")
        
        # Print summary statistics
        total_tokens = sum(speech.get('token_count', 0) for speech in combined_data)
        print(f"Total token count across all speeches: {total_tokens:,}")
        
        # Print breakdown by speaker/category
        speakers = {}
        categories = {}
        
        for speech in combined_data:
            # Count by speaker
            speaker = speech.get('speaker', 'Unknown')
            speakers[speaker] = speakers.get(speaker, 0) + 1
            
            # Count by category
            category = speech.get('category', 'Unknown')
            categories[category] = categories.get(category, 0) + 1
        
        print("\nTop speakers:")
        for speaker, count in sorted(speakers.items(), key=lambda x: x[1], reverse=True)[:10]:
            print(f"  {speaker}: {count}")
        
        print("\nCategories:")
        for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
            print(f"  {category}: {count}")
    
    def run(self):
        """Main scraping function"""
        print("Starting BIS Central Bank Speeches Scraper")
        print(f"Max pages to scrape: {MAX_PAGES if MAX_PAGES else 'All'}")
        print(f"Using Selenium: {self.use_selenium}")
        print(f"Output file: {OUTPUT_FILE}")
        
        try:
            speeches = self.scrape_all_pages()
            self.save_data(speeches)
            print(f"\nScraping completed successfully!")
            
        except KeyboardInterrupt:
            print("Scraping interrupted by user")
        except Exception as e:
            print(f"Scraping failed: {e}")
            raise
        finally:
            if self.driver:
                self.driver.quit()
                print("Selenium driver closed")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='BIS Central Bank Speeches Scraper')
    parser.add_argument('--no-selenium', action='store_true', help='Disable Selenium (use requests only)')
    parser.add_argument('--pages', type=int, help='Maximum pages to scrape (default: 3)')
    parser.add_argument('--full', action='store_true', help='Scrape all pages (overrides --pages)')
    
    args = parser.parse_args()
    
    # Update global MAX_PAGES based on arguments
    global MAX_PAGES
    if args.full:
        MAX_PAGES = None
    elif args.pages:
        MAX_PAGES = args.pages
    
    # Initialize and run scraper
    use_selenium = SELENIUM_AVAILABLE and not args.no_selenium
    scraper = BISSpeechesScraper(use_selenium=use_selenium)
    scraper.run()


if __name__ == "__main__":
    main()