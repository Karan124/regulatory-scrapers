#!/usr/bin/env python3
"""
BIS News & Publications Scraper
Comprehensive scraper for Bank for International Settlements website
Extracts articles, PDFs, Excel files, and metadata with anti-bot evasion
"""

import requests
import json
import os
import time
import re
import zipfile
import io
from datetime import datetime
from urllib.parse import urljoin, urlparse
from pathlib import Path
from typing import Dict, List, Optional, Any
import hashlib

# Third-party imports
from bs4 import BeautifulSoup
import pandas as pd
from fake_useragent import UserAgent
import fitz  # PyMuPDF for PDF extraction
import openpyxl
from openpyxl import load_workbook
import tiktoken

# Optional Selenium imports for JavaScript handling
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("Selenium not available. Install with: pip install selenium webdriver-manager")

# Configuration
BASE_URL = "https://www.bis.org"
START_URL = "https://www.bis.org/press/wnew.htm?m=257"
DATA_DIR = Path("data")
OUTPUT_FILE = DATA_DIR / "bis_news.json"

# Set MAX_PAGE for different run types
MAX_PAGES = 3  # Change to None for full scrape, or set to desired number

class BISScraper:
    def __init__(self, use_selenium: bool = True):
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.session = requests.Session()
        self.ua = UserAgent()
        self.driver = None
        self.setup_session()
        self.existing_urls = set()
        self.encoding = tiktoken.get_encoding("cl100k_base")  # GPT-4 tokenizer
        
        # Create data directory
        DATA_DIR.mkdir(exist_ok=True)
        
        # Load existing data to avoid duplicates
        self.load_existing_data()
        
        if self.use_selenium:
            self.setup_selenium()
    
    def setup_session(self):
        """Setup session with realistic headers and cookies"""
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0'
        })
        
        # Visit root page first to collect cookies
        try:
            response = self.session.get(BASE_URL, timeout=10)
            print(f"Root page status: {response.status_code}")
            time.sleep(2)
        except Exception as e:
            print(f"Warning: Could not visit root page: {e}")
    
    def setup_selenium(self):
        """Setup Selenium WebDriver for JavaScript handling"""
        if not SELENIUM_AVAILABLE:
            return
        
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument(f'--user-agent={self.ua.random}')
        
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            print("Selenium WebDriver initialized successfully")
        except Exception as e:
            print(f"Failed to initialize Selenium: {e}")
            self.use_selenium = False
    
    def get_page_content_selenium(self, url: str) -> Optional[str]:
        """Get page content using Selenium for JavaScript sites"""
        if not self.driver:
            return None
        
        try:
            self.driver.get(url)
            
            # Wait for the content to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "documentList"))
            )
            
            time.sleep(3)  # Additional wait for dynamic content
            return self.driver.page_source
            
        except Exception as e:
            print(f"Selenium error for {url}: {e}")
            return None
    def load_existing_data(self):
        """Load existing scraped data to avoid duplicates"""
        if OUTPUT_FILE.exists():
            try:
                with open(OUTPUT_FILE, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                    self.existing_urls = {item.get('url', '') for item in existing_data}
                print(f"Loaded {len(self.existing_urls)} existing articles")
            except Exception as e:
                print(f"Warning: Could not load existing data: {e}")
                self.existing_urls = set()
    
    def count_tokens(self, text: str) -> int:
        """Count tokens using tiktoken encoding"""
        if not text:
            return 0
        try:
            return len(self.encoding.encode(text))
        except Exception:
            # Fallback to simple word count * 1.3 approximation
            return int(len(text.split()) * 1.3)
    
    def safe_request(self, url: str, retries: int = 3) -> Optional[requests.Response]:
        """Make a safe request with retries and error handling"""
        for attempt in range(retries):
            try:
                # Update headers for each request
                self.session.headers.update({
                    'Referer': BASE_URL,
                    'Sec-Fetch-Site': 'same-origin' if BASE_URL in url else 'cross-site'
                })
                
                response = self.session.get(url, timeout=15)
                if response.status_code == 200:
                    return response
                elif response.status_code == 429:
                    wait_time = (attempt + 1) * 5
                    print(f"Rate limited. Waiting {wait_time}s...")
                    time.sleep(wait_time)
                else:
                    print(f"HTTP {response.status_code} for {url}")
                    
            except Exception as e:
                print(f"Request error (attempt {attempt + 1}): {e}")
                if attempt < retries - 1:
                    time.sleep((attempt + 1) * 2)
        
        return None
    
    def extract_pdf_text(self, pdf_url: str) -> str:
        """Extract text from PDF files"""
        try:
            response = self.safe_request(pdf_url)
            if not response:
                return ""
            
            pdf_document = fitz.open(stream=response.content, filetype="pdf")
            text_content = []
            
            for page_num in range(len(pdf_document)):
                page = pdf_document.load_page(page_num)
                text_content.append(page.get_text())
            
            pdf_document.close()
            return "\n\n".join(text_content)
            
        except Exception as e:
            print(f"PDF extraction error for {pdf_url}: {e}")
            return ""
    
    def extract_excel_data(self, excel_url: str) -> str:
        """Extract data from Excel files"""
        try:
            response = self.safe_request(excel_url)
            if not response:
                return ""
            
            # Try to read with pandas first
            try:
                excel_file = pd.ExcelFile(io.BytesIO(response.content))
                all_data = []
                
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    all_data.append(f"Sheet: {sheet_name}\n{df.to_string()}")
                
                return "\n\n".join(all_data)
                
            except Exception:
                # Fallback to openpyxl
                workbook = load_workbook(io.BytesIO(response.content), data_only=True)
                all_data = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_data = []
                    
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            sheet_data.append('\t'.join(str(cell) if cell is not None else '' for cell in row))
                    
                    if sheet_data:
                        all_data.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_data))
                
                return "\n\n".join(all_data)
                
        except Exception as e:
            print(f"Excel extraction error for {excel_url}: {e}")
            return ""
    
    def extract_zip_contents(self, zip_url: str) -> str:
        """Extract text content from ZIP files"""
        try:
            response = self.safe_request(zip_url)
            if not response:
                return ""
            
            with zipfile.ZipFile(io.BytesIO(response.content)) as zip_file:
                extracted_content = []
                
                for file_info in zip_file.infolist():
                    if file_info.filename.lower().endswith(('.txt', '.csv', '.json')):
                        try:
                            content = zip_file.read(file_info).decode('utf-8', errors='ignore')
                            extracted_content.append(f"File: {file_info.filename}\n{content}")
                        except Exception as e:
                            print(f"Error extracting {file_info.filename}: {e}")
                
                return "\n\n".join(extracted_content)
                
        except Exception as e:
            print(f"ZIP extraction error for {zip_url}: {e}")
            return ""
    
    def extract_tables(self, soup: BeautifulSoup) -> str:
        """Extract table data from soup"""
        tables = soup.find_all('table')
        table_data = []
        
        for i, table in enumerate(tables):
            rows = []
            for row in table.find_all('tr'):
                cells = []
                for cell in row.find_all(['td', 'th']):
                    cell_text = cell.get_text(strip=True)
                    cells.append(cell_text)
                if cells:
                    rows.append('\t'.join(cells))
            
            if rows:
                table_data.append(f"Table {i+1}:\n" + '\n'.join(rows))
        
        return "\n\n".join(table_data)
    
    def extract_related_links(self, soup: BeautifulSoup, base_url: str) -> List[str]:
        """Extract related links from article content"""
        content_div = soup.find('div', id='cmsContent')
        if not content_div:
            content_div = soup
        
        links = []
        for link in content_div.find_all('a', href=True):
            href = link['href']
            
            # Skip irrelevant links
            if any(skip in href.lower() for skip in ['mailto:', 'javascript:', '#', 'about', 'contact', 'terms']):
                continue
            
            # Convert relative URLs to absolute
            full_url = urljoin(base_url, href)
            
            # Only include relevant links
            if any(domain in full_url for domain in ['bis.org', 'fsb.org', 'iosco.org']):
                links.append(full_url)
        
        return list(set(links))  # Remove duplicates
    
    def parse_article_from_index(self, article_element) -> Dict[str, Any]:
        """Parse article metadata from index page"""
        try:
            # Extract date
            date_elem = article_element.find('td', class_='item_date')
            published_date = date_elem.get_text(strip=True) if date_elem else ""
            
            # Extract title and URL
            title_link = article_element.find('a', class_='dark')
            if not title_link:
                return None
            
            headline = title_link.get_text(strip=True)
            url = urljoin(BASE_URL, title_link['href'])
            
            # Extract category
            category_elem = article_element.find('span')
            category = category_elem.get_text(strip=True) if category_elem else ""
            
            # Extract authors
            authors = []
            author_links = article_element.find_all('a', class_='authorlnk')
            for author_link in author_links:
                authors.append(author_link.get_text(strip=True))
            
            return {
                'headline': headline,
                'url': url,
                'published_date': published_date,
                'category': category,
                'authors': authors
            }
            
        except Exception as e:
            print(f"Error parsing article element: {e}")
            return None
    
    def scrape_article_content(self, article_data: Dict[str, Any]) -> Dict[str, Any]:
        """Scrape full content from article page or direct PDF"""
        url = article_data['url']
        
        # Skip if already scraped
        if url in self.existing_urls:
            print(f"Skipping already scraped: {article_data['headline']}")
            return None
        
        print(f"Scraping: {article_data['headline']}")
        
        # Check if this is a direct PDF link
        if url.lower().endswith('.pdf'):
            print(f"  Direct PDF link detected: {url}")
            
            # Extract PDF content directly
            pdf_text = self.extract_pdf_text(url)
            
            # Calculate token count
            token_count = self.count_tokens(pdf_text)
            
            # Build final article data for PDF-only content
            final_data = {
                **article_data,
                'scraped_date': datetime.now().isoformat(),
                'content': f"PDF Document: {article_data['headline']}\n\nThis content is available as a PDF document.",
                'token_count': token_count,
                'related_links': [],
                'associated_image': "",
                'attachments': {
                    'pdf_text': pdf_text,
                    'excel_data': '',
                    'tables': '',
                    'charts': ''
                }
            }
            
            return final_data
        
        # Regular web page processing
        response = self.safe_request(url)
        if not response:
            return None
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extract main content
        content_div = soup.find('div', {'id': 'center', 'role': 'main'})
        if not content_div:
            content_div = soup.find('div', id='cmsContent')
        
        main_content = ""
        if content_div:
            # Remove child documents section
            child_docs = content_div.find('div', class_='childdocsgroup')
            if child_docs:
                child_docs.decompose()
            
            main_content = content_div.get_text(separator='\n', strip=True)
        
        # Extract image
        img_elem = soup.find('img', src=True)
        associated_image = ""
        if img_elem and 'src' in img_elem.attrs:
            associated_image = urljoin(BASE_URL, img_elem['src'])
        
        # Extract tables
        tables_text = self.extract_tables(soup)
        
        # Extract related links
        related_links = self.extract_related_links(soup, url)
        
        # Process attachments
        attachments = {
            'pdf_text': '',
            'excel_data': '',
            'tables': tables_text,
            'charts': ''
        }
        
        # Find PDF, Excel, and other file links
        file_links = soup.find_all('a', href=True)
        for link in file_links:
            href = link['href']
            if not href:
                continue
            
            file_url = urljoin(BASE_URL, href)
            
            if href.lower().endswith('.pdf'):
                print(f"  Extracting PDF: {file_url}")
                attachments['pdf_text'] += self.extract_pdf_text(file_url) + "\n\n"
                time.sleep(1)  # Be respectful
                
            elif href.lower().endswith(('.xlsx', '.xls')):
                print(f"  Extracting Excel: {file_url}")
                attachments['excel_data'] += self.extract_excel_data(file_url) + "\n\n"
                time.sleep(1)
                
            elif href.lower().endswith('.zip'):
                print(f"  Extracting ZIP: {file_url}")
                attachments['excel_data'] += self.extract_zip_contents(file_url) + "\n\n"
                time.sleep(1)
        
        # Calculate total token count
        all_text = main_content + attachments['pdf_text'] + attachments['excel_data'] + attachments['tables']
        token_count = self.count_tokens(all_text)
        
        # Build final article data
        final_data = {
            **article_data,
            'scraped_date': datetime.now().isoformat(),
            'content': main_content,
            'token_count': token_count,
            'related_links': related_links,
            'associated_image': associated_image,
            'attachments': attachments
        }
        
        return final_data
    
    def get_articles_from_page(self, url: str) -> List[Dict[str, Any]]:
        """Extract all articles from an index page"""
        # Try Selenium first if available
        html_content = None
        if self.use_selenium:
            print("Using Selenium to get page content...")
            html_content = self.get_page_content_selenium(url)
        
        # Fallback to requests if Selenium fails or not available
        if not html_content:
            print("Using requests to get page content...")
            response = self.safe_request(url)
            if not response:
                return []
            html_content = response.text
        
        soup = BeautifulSoup(html_content, 'html.parser')
        articles = []
        
        # Find article rows - the ORIGINAL working approach
        article_rows = soup.find_all('tr', class_='item even') + soup.find_all('tr', class_='item odd')
        
        print(f"Found {len(article_rows)} article rows in HTML")
        
        for row in article_rows:
            article_data = self.parse_article_from_index(row)
            if article_data:
                articles.append(article_data)
        
        return articles
    
    def has_next_page(self, soup: BeautifulSoup, current_page: int) -> Optional[str]:
        """Check if there's a next page and return its URL"""
        # Look for next button
        next_button = soup.find('a', class_='navcarot')
        if next_button and 'href' in next_button.attrs:
            return urljoin(BASE_URL, next_button['href'])
        
        # Fallback: construct next page URL manually
        base_params = "?m=257"
        if current_page == 1:
            next_url = f"{START_URL.split('?')[0]}{base_params}&newsarchive_page=2"
        else:
            next_url = f"{START_URL.split('?')[0]}{base_params}&newsarchive_page={current_page + 1}"
        
        # Check if we're at the last page by looking for page info
        page_info = soup.find('div', class_='pageof')
        if page_info:
            page_text = page_info.get_text()
            # Extract "Page X of Y" info
            match = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', page_text)
            if match:
                current = int(match.group(1))
                total = int(match.group(2))
                print(f"Page {current} of {total}")
                if current >= total:
                    return None  # Last page
        
        return next_url
    
    def scrape_all_pages(self) -> List[Dict[str, Any]]:
        """Scrape articles from all pages"""
        all_articles = []
        current_url = START_URL
        page_num = 1
        
        while current_url and (MAX_PAGES is None or page_num <= MAX_PAGES):
            print(f"\n=== Scraping Page {page_num} ===")
            print(f"URL: {current_url}")
            
            # Get articles from current page
            articles = self.get_articles_from_page(current_url)
            print(f"Found {len(articles)} articles on page {page_num}")
            
            if not articles:
                print("No articles found, stopping.")
                break
            
            # Scrape full content for each article
            for article in articles:
                full_article = self.scrape_article_content(article)
                if full_article:
                    all_articles.append(full_article)
                
                time.sleep(2)  # Be respectful to the server
            
            # Find next page
            response = self.safe_request(current_url)
            if response:
                soup = BeautifulSoup(response.content, 'html.parser')
                next_url = self.has_next_page(soup, page_num)
                
                if next_url:
                    current_url = next_url
                    page_num += 1
                    time.sleep(3)  # Pause between pages
                else:
                    print("No more pages found.")
                    break
            else:
                break
        
        return all_articles
    
    def save_data(self, articles: List[Dict[str, Any]]):
        """Save scraped data to JSON file"""
        # Load existing data
        existing_data = []
        if OUTPUT_FILE.exists():
            try:
                with open(OUTPUT_FILE, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except Exception as e:
                print(f"Error loading existing data: {e}")
        
        # Merge new articles (avoid duplicates by URL)
        existing_urls = {item.get('url') for item in existing_data}
        new_articles = [article for article in articles if article.get('url') not in existing_urls]
        
        combined_data = existing_data + new_articles
        
        # Save to file
        with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
            json.dump(combined_data, f, indent=2, ensure_ascii=False)
        
        print(f"\nSaved {len(new_articles)} new articles ({len(combined_data)} total)")
        
        # Print summary statistics
        total_tokens = sum(article.get('token_count', 0) for article in combined_data)
        print(f"Total token count across all articles: {total_tokens:,}")
    
    def run(self):
        """Main scraping function"""
        print("Starting BIS News & Publications Scraper")
        print(f"Max pages to scrape: {MAX_PAGES if MAX_PAGES else 'All'}")
        print(f"Using Selenium: {self.use_selenium}")
        
        try:
            articles = self.scrape_all_pages()
            self.save_data(articles)
            print(f"\nScraping completed successfully!")
            
        except Exception as e:
            print(f"Scraping failed: {e}")
            raise
        finally:
            if self.driver:
                self.driver.quit()


def main():
    """Main entry point"""
    # Try with Selenium first if available, then fallback to requests
    use_selenium = SELENIUM_AVAILABLE
    
    scraper = BISScraper(use_selenium=use_selenium)
    scraper.run()


if __name__ == "__main__":
    main()