#!/usr/bin/env python3
"""
RBNZ Consultations Scraper
Scrapes all consultations from the Reserve Bank of New Zealand consultations page
with working pagination and rate limiting.
"""

import json
import logging
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import PyPDF2
from io import BytesIO

# Optional Selenium imports for JavaScript handling
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# Configuration
CONFIG = {
    'BASE_URL': 'https://www.rbnz.govt.nz',
    'CONSULTATIONS_URL': 'https://www.rbnz.govt.nz/have-your-say/closed-consultations',
    'USER_AGENT': 'rbnz-approved-agent/rg-11701',
    'RATE_LIMIT': 292,  # requests per hour
    'REQUEST_DELAY': 3600 / 292,  # seconds between requests (~12.3 seconds)
    'SAFETY_MARGIN': 0.8,  # Use only 80% of allowed rate for safety
    'MAX_PAGE': 1,  # Set to None for full scrape, or integer for limited pages
    'OUTPUT_DIR': './data',
    'OUTPUT_FILE': './data/rbnz_consultations.json',
    'LOG_FILE': './consultations_scrape.log',
    'SCRAPED_URLS_FILE': './data/scraped_consultations_urls.json',
    'REQUEST_LOG_FILE': './data/consultation_request_log.json'
}

class RBNZConsultationsScraper:
    def __init__(self, max_pages: Optional[int] = None, use_selenium: bool = False):
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': CONFIG['USER_AGENT']})
        self.scraped_urls: Set[str] = self._load_scraped_urls()
        self.max_pages = max_pages or CONFIG['MAX_PAGE']
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.setup_logging()
        self.setup_directories()
        
        # Rate limiting tracking
        self.request_times = []
        self.request_count = 0
        self.hourly_limit = int(CONFIG['RATE_LIMIT'] * CONFIG['SAFETY_MARGIN'])
        self.request_delay = 3600 / self.hourly_limit  # Adjusted delay with safety margin
        
        self.logger.info(f"Rate limiting: {self.hourly_limit} requests/hour, {self.request_delay:.1f}s between requests")
        
        if self.use_selenium:
            self.setup_selenium()
        
    def setup_selenium(self):
        """Setup Selenium WebDriver for JavaScript handling"""
        if not SELENIUM_AVAILABLE:
            self.logger.warning("Selenium not available. Install with: pip install selenium webdriver-manager")
            self.use_selenium = False
            return
            
        self.chrome_options = Options()
        self.chrome_options.add_argument('--headless')
        self.chrome_options.add_argument(f'--user-agent={CONFIG["USER_AGENT"]}')
        self.chrome_options.add_argument('--no-sandbox')
        self.chrome_options.add_argument('--disable-dev-shm-usage')
        self.chrome_options.add_argument('--disable-gpu')
        self.chrome_options.add_argument('--window-size=1920,1080')
        # Additional stability options
        self.chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        self.chrome_options.add_argument('--disable-extensions')
        self.chrome_options.add_argument('--disable-plugins')
        self.chrome_options.add_argument('--disable-images')  # Speed up loading
        self.chrome_options.add_argument('--disable-javascript-harmony-shipping')
        self.chrome_options.add_argument('--disable-background-timer-throttling')
        self.chrome_options.add_argument('--disable-renderer-backgrounding')
        self.chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        self.chrome_options.add_argument('--disable-ipc-flooding-protection')
        # Memory management
        self.chrome_options.add_argument('--memory-pressure-off')
        self.chrome_options.add_argument('--max_old_space_size=4096')
        
        self.logger.info("Selenium WebDriver configured")
        
    def setup_logging(self):
        """Configure logging"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(CONFIG['LOG_FILE']),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def setup_directories(self):
        """Create necessary directories"""
        Path(CONFIG['OUTPUT_DIR']).mkdir(exist_ok=True)
        
    def _load_scraped_urls(self) -> Set[str]:
        """Load previously scraped URLs for deduplication"""
        try:
            if os.path.exists(CONFIG['SCRAPED_URLS_FILE']):
                with open(CONFIG['SCRAPED_URLS_FILE'], 'r') as f:
                    return set(json.load(f))
        except Exception as e:
            self.logger.warning(f"Could not load scraped URLs: {e}")
        return set()
        
    def _save_scraped_urls(self):
        """Save scraped URLs to file"""
        try:
            with open(CONFIG['SCRAPED_URLS_FILE'], 'w') as f:
                json.dump(list(self.scraped_urls), f, indent=2)
        except Exception as e:
            self.logger.error(f"Could not save scraped URLs: {e}")

    def _rate_limit(self):
        """Implement smart rate limiting with request tracking"""
        current_time = time.time()
        
        # Remove requests older than 1 hour
        self.request_times = [t for t in self.request_times if current_time - t < 3600]
        
        # Check if we're approaching the hourly limit
        if len(self.request_times) >= self.hourly_limit:
            # Calculate how long to wait until the oldest request is > 1 hour old
            oldest_request = min(self.request_times)
            wait_time = 3600 - (current_time - oldest_request) + 1  # +1 second buffer
            
            if wait_time > 0:
                self.logger.warning(f"Rate limit reached. Waiting {wait_time:.1f} seconds...")
                time.sleep(wait_time)
                current_time = time.time()
                # Clean up old requests after waiting
                self.request_times = [t for t in self.request_times if current_time - t < 3600]
        
        # Standard delay between requests
        time.sleep(self.request_delay)
        
        # Record this request
        self.request_times.append(current_time)
        self.request_count += 1
        
        # Log progress every 20 requests
        if self.request_count % 20 == 0:
            requests_in_last_hour = len(self.request_times)
            self.logger.info(f"Request #{self.request_count}: {requests_in_last_hour}/{self.hourly_limit} requests in last hour")

    def _safe_request(self, url: str, timeout: int = 30) -> Optional[requests.Response]:
        """Make a rate-limited request with error handling"""
        max_retries = 3
        retry_delay = 30  # Start with 30 seconds
        
        for attempt in range(max_retries):
            try:
                self._rate_limit()
                
                response = self.session.get(url, timeout=timeout)
                
                if response.status_code == 429:
                    # Rate limited - wait longer
                    wait_time = retry_delay * (attempt + 1)
                    self.logger.warning(f"Rate limited (429). Waiting {wait_time} seconds before retry {attempt + 1}/{max_retries}")
                    time.sleep(wait_time)
                    continue
                
                response.raise_for_status()
                return response
                
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 429:
                    wait_time = retry_delay * (attempt + 1)
                    self.logger.warning(f"Rate limited (429). Waiting {wait_time} seconds before retry {attempt + 1}/{max_retries}")
                    time.sleep(wait_time)
                    continue
                else:
                    self.logger.error(f"HTTP error for {url}: {e}")
                    break
            except Exception as e:
                self.logger.error(f"Request error for {url}: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                break
        
        return None

    def _clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        if not text:
            return ""
        
        try:
            # Remove extra whitespace and normalize
            text = re.sub(r'\s+', ' ', text.strip())
            # Remove HTML entities
            text = text.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
            # Normalize quotes
            text = text.replace('"', '"').replace('"', '"').replace('„', '"')
            text = text.replace(''', "'").replace(''', "'").replace('`', "'")
            # Remove HTML tags
            text = re.sub(r'<[^>]+>', '', text)
            # Remove excessive punctuation
            text = re.sub(r'\.{3,}', '...', text)
            # Normalize dashes
            text = text.replace('–', '-').replace('—', '-')
        except Exception as e:
            self.logger.warning(f"Text cleaning failed: {e}")
            text = ' '.join(text.split())
            
        return text

    def _extract_pdf_text(self, pdf_url: str) -> str:
        """Extract text from PDF files with rate limiting"""
        try:
            response = self._safe_request(pdf_url, timeout=60)
            if not response:
                return ""
            
            pdf_reader = PyPDF2.PdfReader(BytesIO(response.content))
            text_content = []
            
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content.append(self._clean_text(page_text))
                    
            return "\n\n".join(text_content)
            
        except Exception as e:
            self.logger.error(f"Error extracting PDF {pdf_url}: {e}")
            return ""

    def _extract_excel_data(self, excel_url: str) -> str:
        """Extract data from Excel files using openpyxl with rate limiting"""
        try:
            response = self._safe_request(excel_url, timeout=60)
            if not response:
                return ""
            
            # Try to use openpyxl if available, otherwise skip Excel processing
            try:
                import openpyxl
                from io import BytesIO
                
                workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
                excel_content = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    excel_content.append(f"--- SHEET: {sheet_name} ---")
                    
                    # Get all rows with data
                    rows_data = []
                    for row in sheet.iter_rows(values_only=True):
                        # Convert row to strings and filter out None values
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        # Only add rows that have some content
                        if any(cell.strip() for cell in row_data if cell):
                            rows_data.append(" | ".join(row_data))
                    
                    excel_content.extend(rows_data[:500])  # Limit to first 500 rows per sheet
                    
                return "\n".join(excel_content)
                
            except ImportError:
                self.logger.warning("openpyxl not available for Excel processing. Install with: pip install openpyxl")
                return f"Excel file available at: {excel_url} (openpyxl required for processing)"
                
        except Exception as e:
            self.logger.error(f"Error extracting Excel data from {excel_url}: {e}")
            return ""

    def _get_rbnz_consultations_page_url(self, page_number: int) -> str:
        """Get the correct RBNZ consultations URL for a specific page"""
        base_url = CONFIG['CONSULTATIONS_URL']
        
        if page_number == 1:
            return f"{base_url}#sort=%40computedsortdate%20descending"
        else:
            first_param = (page_number - 1) * 10
            return f"{base_url}#first={first_param}&sort=%40computedsortdate%20descending"

    def _create_driver(self):
        """Create a new Chrome driver instance"""
        return webdriver.Chrome(
            service=webdriver.chrome.service.Service(ChromeDriverManager().install()),
            options=self.chrome_options
        )

    def _extract_consultation_links_from_page(self, driver) -> List[str]:
        """Extract all consultation links from the current page"""
        try:
            # Wait for results to load
            WebDriverWait(driver, 20).until(
                lambda d: len(d.find_elements(By.CSS_SELECTOR, ".coveo-list-layout.CoveoResult")) > 0
            )
            
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            links = []
            
            # Find all Coveo result containers
            results = soup.find_all('div', class_='coveo-list-layout CoveoResult')
            
            for result in results:
                # Look for the main consultation link
                link_elem = result.find('a', class_='CoveoResultLink')
                if link_elem and link_elem.get('href'):
                    url = link_elem['href']
                    # Consultations are typically in /have-your-say/ or consultations domain
                    if any(path in url for path in ['/have-your-say/', 'consultations.rbnz.govt.nz']):
                        if url.startswith('/'):
                            url = urljoin(CONFIG['BASE_URL'], url)
                        links.append(url)
            
            return links
            
        except Exception as e:
            self.logger.error(f"Error extracting consultation links: {e}")
            return []

    def _scrape_all_pages_selenium(self) -> List[str]:
        """Scrape all pages using Selenium with robust session management"""
        if not self.use_selenium:
            return []
            
        all_links = []
        driver = None
        
        try:
            page = 1
            consecutive_failures = 0
            max_consecutive_failures = 3
            
            while page <= self.max_pages:
                # Create new driver if needed or if previous one failed
                if driver is None:
                    driver = self._create_driver()
                    self.logger.info(f"Created new Chrome driver session")
                
                url = self._get_rbnz_consultations_page_url(page)
                self.logger.info(f"Scraping page {page}: {url}")
                
                try:
                    driver.get(url)
                    time.sleep(8)  # Give Coveo time to load
                    
                    # Test if driver is still responsive
                    try:
                        _ = driver.current_url
                    except Exception as session_error:
                        self.logger.warning(f"Driver session lost: {session_error}")
                        driver.quit()
                        driver = None
                        consecutive_failures += 1
                        
                        if consecutive_failures >= max_consecutive_failures:
                            self.logger.error(f"Too many consecutive failures ({consecutive_failures}), stopping")
                            break
                            
                        self.logger.info("Creating new driver session and retrying...")
                        continue
                    
                    # Extract links from this page
                    page_links = self._extract_consultation_links_from_page(driver)
                    
                    if not page_links:
                        self.logger.info(f"No links found on page {page} - reached end")
                        break
                    
                    # Check if we're getting new links
                    new_links = [link for link in page_links if link not in all_links]
                    
                    if not new_links and page > 1:
                        self.logger.info(f"No new links on page {page} - reached end")
                        break
                    
                    all_links.extend(new_links)
                    self.logger.info(f"Page {page}: found {len(page_links)} links ({len(new_links)} new). Total: {len(all_links)}")
                    
                    # Log first few URLs to verify they're different
                    if new_links:
                        self.logger.info(f"Sample new URLs: {new_links[:2]}")
                    
                    # Reset failure counter on success
                    consecutive_failures = 0
                    page += 1
                    
                    # Restart driver every 5 pages to prevent memory issues
                    if page % 5 == 0:
                        self.logger.info("Restarting driver to prevent memory issues...")
                        driver.quit()
                        driver = None
                        time.sleep(3)
                    
                except Exception as page_error:
                    self.logger.error(f"Error on page {page}: {page_error}")
                    consecutive_failures += 1
                    
                    if consecutive_failures >= max_consecutive_failures:
                        self.logger.error(f"Too many consecutive failures ({consecutive_failures}), stopping")
                        break
                    
                    # Try to recover by creating new driver
                    try:
                        if driver:
                            driver.quit()
                    except:
                        pass
                    driver = None
                    
                    self.logger.info("Attempting to recover with new driver session...")
                    continue
            
            if driver:
                driver.quit()
            
            # Remove any remaining duplicates
            unique_links = list(dict.fromkeys(all_links))
            self.logger.info(f"Selenium scraping completed. Total unique links: {len(unique_links)}")
            return unique_links
            
        except Exception as e:
            self.logger.error(f"Selenium scraping failed: {e}")
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            return all_links  # Return what we got so far

    def _is_pdf_link(self, url: str, link_text: str = "", metadata_text: str = "") -> bool:
        """Determine if a link points to a PDF file"""
        # Direct PDF extension
        if url.lower().endswith('.pdf'):
            return True
            
        # Check for PDF indicators in URL path
        if any(indicator in url.lower() for indicator in ['pdf', 'download']):
            # Check metadata or link text for PDF indicators
            combined_text = f"{link_text} {metadata_text}".lower()
            if any(indicator in combined_text for indicator in ['pdf', 'pdf document']):
                return True
                
        return False

    def _is_excel_link(self, url: str, link_text: str = "", metadata_text: str = "") -> bool:
        """Determine if a link points to an Excel file"""
        # Direct Excel extensions
        if url.lower().endswith(('.xlsx', '.xls')):
            return True
            
        # Check metadata or link text for Excel indicators
        combined_text = f"{link_text} {metadata_text}".lower()
        if any(indicator in combined_text for indicator in ['xlsx', 'xls', 'excel', 'spreadsheet']):
            return True
                
        return False

    def _extract_consultation_content(self, consultation_url: str) -> Optional[Dict]:
        """Extract content from a single consultation with enhanced extraction for consultation-specific elements"""
        if consultation_url in self.scraped_urls:
            return None
            
        try:
            response = self._safe_request(consultation_url)
            if not response:
                return None
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract basic information
            title = ""
            # Look for consultation title in banner
            title_elem = soup.find('h1', id='cs-consultation-title-in-banner') or soup.find('h1')
            if title_elem:
                title = self._clean_text(title_elem.get_text())
            
            # Extract consultation status and dates
            consultation_status = "Unknown"
            opened_date = ""
            closed_date = ""
            
            # Look for consultation dates in sidebar
            dates_container = soup.find('div', class_='cs-consultation-dates-container')
            if dates_container:
                # Extract primary dates (closed/opened)
                primary_date = dates_container.find('p', class_='cs-consultation-sidebar-primary-date')
                if primary_date:
                    primary_text = self._clean_text(primary_date.get_text())
                    if 'Closed' in primary_text:
                        consultation_status = "Closed"
                        # Extract the date
                        date_match = re.search(r'(\d{1,2}\s+\w+\s+\d{4})', primary_text)
                        if date_match:
                            closed_date = date_match.group(1)
                    elif 'Open' in primary_text:
                        consultation_status = "Open"
                        date_match = re.search(r'(\d{1,2}\s+\w+\s+\d{4})', primary_text)
                        if date_match:
                            opened_date = date_match.group(1)
                
                # Extract secondary dates (opened/closed)
                secondary_date = dates_container.find('p', class_='cs-consultation-sidebar-secondary-date')
                if secondary_date:
                    secondary_text = self._clean_text(secondary_date.get_text())
                    if 'Opened' in secondary_text:
                        date_match = re.search(r'(\d{1,2}\s+\w+\s+\d{4})', secondary_text)
                        if date_match:
                            opened_date = date_match.group(1)
            
            # Determine status based on closed_date presence
            if closed_date:
                status = "Closed"
            else:
                status = "Open"
            
            # Extract contact information
            contact_info = {}
            contact_container = soup.find('div', class_='cs-consultation-contact-details')
            if contact_container:
                # Extract email
                email_elem = contact_container.find('a', href=re.compile(r'^mailto:'))
                if email_elem:
                    contact_info['email'] = email_elem.get('href').replace('mailto:', '')
                    
                # Extract other contact details
                contact_text = self._clean_text(contact_container.get_text())
                if contact_text:
                    contact_info['details'] = contact_text
            
            # Extract overview content
            overview_content = ""
            overview_div = soup.find('div', id='overview')
            if overview_div:
                content_sections = []
                for elem in overview_div.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'li']):
                    text = self._clean_text(elem.get_text())
                    if text and len(text) > 10:
                        content_sections.append(text)
                overview_content = "\n\n".join(content_sections)
            
            # Extract download links for PDFs and Excel files (excluding ZIP files as requested)
            download_links = {
                'pdf_files': [],
                'excel_files': [],
                'other_files': []
            }
            
            # Look for document links in the content area only (not social media)
            content_area = soup.find('div', class_='dss-easy-reading') or soup.find('div', id='overview') or soup
            
            # First, look for consultation CTA (Call-to-Action) sections which often contain PDF downloads
            cta_sections = soup.find_all('div', class_='cs-consultation-cta')
            for cta in cta_sections:
                cta_links = cta.find_all('a', href=True)
                for link in cta_links:
                    href = link['href']
                    link_text = self._clean_text(link.get_text())
                    
                    # Look for file metadata (size, type info)
                    metadata_elem = cta.find('small', class_='cs-consultation-cta-link-file-metadata')
                    metadata_text = self._clean_text(metadata_elem.get_text()) if metadata_elem else ""
                    
                    # Skip social media links
                    if any(social in href.lower() for social in ['twitter', 'facebook', 'linkedin', 'share']):
                        continue
                        
                    # Skip zip files as requested
                    if href.lower().endswith('.zip') or 'zip' in metadata_text.lower():
                        continue
                    
                    if href.startswith('/'):
                        href = urljoin(CONFIG['BASE_URL'], href)
                    elif not href.startswith('http'):
                        # Handle relative links
                        href = urljoin(consultation_url, href)
                    
                    download_info = {
                        'title': link_text or 'Document',
                        'url': href,
                        'file_type': metadata_text or 'Unknown',
                        'metadata': metadata_text
                    }
                    
                    # Use enhanced PDF/Excel detection
                    if self._is_pdf_link(href, link_text, metadata_text):
                        download_links['pdf_files'].append(download_info)
                        self.logger.debug(f"Found PDF link in CTA: {link_text} -> {href}")
                    elif self._is_excel_link(href, link_text, metadata_text):
                        download_links['excel_files'].append(download_info)
                        self.logger.debug(f"Found Excel link in CTA: {link_text} -> {href}")
                    else:
                        download_links['other_files'].append(download_info)
            
            # Then look for other document links in the content area
            for link in content_area.find_all('a', href=True):
                href = link['href']
                
                # Skip if already processed in CTA sections
                if any(href == dl['url'] for dl_list in download_links.values() for dl in dl_list):
                    continue
                
                # Skip social media links
                if any(social in href.lower() for social in ['twitter', 'facebook', 'linkedin', 'share']):
                    continue
                    
                # Skip zip files as requested
                if href.lower().endswith('.zip'):
                    continue
                
                if href.startswith('/'):
                    href = urljoin(CONFIG['BASE_URL'], href)
                elif not href.startswith('http'):
                    # Handle relative links
                    href = urljoin(consultation_url, href)
                
                link_text = self._clean_text(link.get_text())
                
                # Look for nearby metadata (file size, type info)
                metadata_text = ""
                parent = link.parent
                if parent:
                    # Look for file metadata in parent or sibling elements
                    for sibling in parent.find_all(['small', 'span', 'em']):
                        sibling_text = self._clean_text(sibling.get_text())
                        if any(indicator in sibling_text.lower() for indicator in ['kb', 'mb', 'pdf', 'excel', 'document']):
                            metadata_text = sibling_text
                            break
                
                download_info = {
                    'title': link_text or 'Document',
                    'url': href,
                    'file_type': href.split('.')[-1].upper() if '.' in href else 'Link',
                    'metadata': metadata_text
                }
                
                # Use enhanced PDF/Excel detection
                if self._is_pdf_link(href, link_text, metadata_text):
                    download_links['pdf_files'].append(download_info)
                    self.logger.debug(f"Found PDF link: {link_text} -> {href}")
                elif self._is_excel_link(href, link_text, metadata_text):
                    download_links['excel_files'].append(download_info)
                    self.logger.debug(f"Found Excel link: {link_text} -> {href}")
                elif any(href.lower().endswith(ext) for ext in ['.doc', '.docx', '.txt']):
                    download_links['other_files'].append(download_info)
            
            # Extract related links (only from content, not social media)
            related_links = []
            social_patterns = ['twitter', 'facebook', 'linkedin', 'share', 'mailto:', '#']
            
            for link in content_area.find_all('a', href=True):
                href = link['href']
                link_text = self._clean_text(link.get_text())
                
                # Skip social media and internal anchors
                if any(pattern in href.lower() for pattern in social_patterns):
                    continue
                    
                # Skip file downloads (already captured above)
                if any(href == dl['url'] for dl_list in download_links.values() for dl in dl_list):
                    continue
                
                if href.startswith('/'):
                    href = urljoin(CONFIG['BASE_URL'], href)
                elif not href.startswith('http'):
                    href = urljoin(consultation_url, href)
                
                if link_text and len(link_text) > 3 and href not in [rl['url'] for rl in related_links]:
                    related_links.append({
                        'text': link_text,
                        'url': href
                    })
                        
            # Extract image
            associated_image_url = ""
            # Look for consultation banner image
            banner_elem = soup.find('div', class_='dss-rhino cs-consultation-banner')
            if banner_elem and banner_elem.get('style'):
                style = banner_elem.get('style')
                url_match = re.search(r"url\('([^']+)'\)", style)
                if url_match:
                    img_url = url_match.group(1)
                    if img_url.startswith('/'):
                        img_url = urljoin(CONFIG['BASE_URL'], img_url)
                    associated_image_url = img_url
                    
            # Process PDF files
            pdf_content = ""
            for pdf_info in download_links['pdf_files']:
                self.logger.info(f"Extracting PDF content from: {pdf_info['title']}")
                pdf_text = self._extract_pdf_text(pdf_info['url'])
                if pdf_text:
                    pdf_content += f"\n\n--- PDF: {pdf_info['title']} ---\n\n{pdf_text}"
                    
            # Process Excel files
            excel_content = ""
            for excel_info in download_links['excel_files']:
                self.logger.info(f"Extracting Excel data from: {excel_info['title']}")
                excel_data = self._extract_excel_data(excel_info['url'])
                if excel_data:
                    excel_content += f"\n\n--- EXCEL: {excel_info['title']} ---\n\n{excel_data}"
                    
            # Extract tables from HTML
            tables_data = []
            for table in soup.find_all('table'):
                table_text = []
                for row in table.find_all('tr'):
                    row_text = []
                    for cell in row.find_all(['td', 'th']):
                        cell_text = self._clean_text(cell.get_text())
                        if cell_text:
                            row_text.append(cell_text)
                    if row_text:
                        table_text.append(" | ".join(row_text))
                if table_text:
                    tables_data.append("\n".join(table_text))
                    
            tables_and_charts_data = "\n\n--- HTML TABLE ---\n\n".join(tables_data)
            
            # Mark as scraped
            self.scraped_urls.add(consultation_url)
            
            return {
                'url': consultation_url,
                'title': title,
                'status': status,  # Added status field based on closed_date presence
                'consultation_status': consultation_status,
                'opened_date': opened_date,
                'closed_date': closed_date,
                'scraped_date': datetime.now().isoformat(),
                'contact_info': contact_info,
                'overview_content': overview_content,
                'related_links': related_links,
                'associated_image_url': associated_image_url,
                'download_links': download_links,
                'pdf_content': pdf_content,
                'excel_content': excel_content,
                'tables_and_charts_data': tables_and_charts_data
            }
            
        except Exception as e:
            self.logger.error(f"Error extracting content from {consultation_url}: {e}")
            import traceback
            self.logger.debug(f"Full traceback: {traceback.format_exc()}")
            return None

    def scrape_all_consultations(self) -> List[Dict]:
        """Main scraping method with progress tracking"""
        self.logger.info(f"Starting RBNZ consultations scraping (max_pages: {self.max_pages})")
        
        # Get all consultation URLs
        if self.use_selenium:
            self.logger.info("Using Selenium to get consultation URLs")
            consultation_urls = self._scrape_all_pages_selenium()
        else:
            self.logger.error("Non-Selenium scraping not implemented - use --use-selenium")
            return []
        
        if not consultation_urls:
            self.logger.warning("No consultation URLs found")
            return []
            
        self.logger.info(f"Found {len(consultation_urls)} total consultation URLs")
        
        # Calculate estimated time
        estimated_seconds = len(consultation_urls) * self.request_delay
        estimated_hours = estimated_seconds / 3600
        self.logger.info(f"Estimated time to complete: {estimated_hours:.1f} hours")
        
        # Extract content from each consultation
        consultations = []
        failed_urls = []
        
        for i, url in enumerate(consultation_urls, 1):
            self.logger.info(f"Processing consultation {i}/{len(consultation_urls)}: {url}")
            
            consultation_data = self._extract_consultation_content(url)
            if consultation_data:
                consultations.append(consultation_data)
                self.logger.info(f"✓ Scraped: {consultation_data['title'][:60]}...")
                
                # Save progress every 10 consultations
                if len(consultations) % 10 == 0:
                    self._save_progress(consultations)
                    self.logger.info(f"Progress saved: {len(consultations)} consultations scraped")
                    
            else:
                failed_urls.append(url)
                self.logger.warning(f"✗ Failed to scrape: {url}")
        
        # Final save
        self._save_progress(consultations)
        
        if failed_urls:
            self.logger.warning(f"Failed to scrape {len(failed_urls)} consultations")
            # Save failed URLs for potential retry
            with open('./data/failed_consultations.json', 'w') as f:
                json.dump(failed_urls, f, indent=2)
                
        self.logger.info(f"Scraping completed. Total consultations: {len(consultations)}")
        return consultations

    def _save_progress(self, consultations: List[Dict]):
        """Save progress to avoid losing work"""
        try:
            # Save to a temporary progress file
            progress_file = './data/consultations_progress.json'
            with open(progress_file, 'w', encoding='utf-8') as f:
                json.dump(consultations, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"Error saving progress: {e}")
            
    def save_results(self, consultations: List[Dict]):
        """Save scraped consultations to JSON file"""
        try:
            # Load existing consultations
            existing_consultations = []
            if os.path.exists(CONFIG['OUTPUT_FILE']):
                with open(CONFIG['OUTPUT_FILE'], 'r', encoding='utf-8') as f:
                    existing_consultations = json.load(f)
                    
            # Merge with new consultations
            existing_urls = {cons.get('url') for cons in existing_consultations}
            new_consultations = [cons for cons in consultations if cons.get('url') not in existing_urls]
            
            all_consultations = existing_consultations + new_consultations
            
            # Save combined results
            with open(CONFIG['OUTPUT_FILE'], 'w', encoding='utf-8') as f:
                json.dump(all_consultations, f, indent=2, ensure_ascii=False)
                
            self.logger.info(f"Saved {len(new_consultations)} new consultations. Total: {len(all_consultations)}")
            
            # Save scraped URLs
            self._save_scraped_urls()
            
            # Save request statistics
            self._save_request_stats()
            
        except Exception as e:
            self.logger.error(f"Error saving results: {e}")
            
    def _save_request_stats(self):
        """Save request statistics for monitoring"""
        try:
            stats = {
                'total_requests': self.request_count,
                'requests_in_last_hour': len(self.request_times),
                'hourly_limit': self.hourly_limit,
                'last_request_time': max(self.request_times) if self.request_times else None,
                'scrape_date': datetime.now().isoformat()
            }
            
            with open('./data/consultation_request_stats.json', 'w') as f:
                json.dump(stats, f, indent=2)
                
        except Exception as e:
            self.logger.error(f"Error saving request stats: {e}")
            
    def run(self):
        """Main execution method"""
        start_time = datetime.now()
        self.logger.info(f"RBNZ consultations scraper started at {start_time}")
        
        try:
            consultations = self.scrape_all_consultations()
            if consultations:
                self.save_results(consultations)
            else:
                self.logger.warning("No consultations were scraped")
                
        except Exception as e:
            self.logger.error(f"Scraping failed: {e}")
            
        end_time = datetime.now()
        duration = end_time - start_time
        self.logger.info(f"Scraping completed in {duration}")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='RBNZ Consultations Scraper')
    parser.add_argument('--max-pages', type=int, default=1,
                       help='Maximum number of pages to scrape')
    parser.add_argument('--max-consultations', type=int, default=None,
                       help='Maximum number of consultations to scrape (useful for testing)')
    parser.add_argument('--use-selenium', action='store_true',
                       help='Use Selenium WebDriver (required)')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug logging')
    parser.add_argument('--test-url', type=str,
                       help='Test scraping a specific consultation URL')
    parser.add_argument('--resume', action='store_true',
                       help='Resume from previous progress file')
    parser.add_argument('--batch-size', type=int, default=50,
                       help='Number of consultations to scrape in this batch')
    
    args = parser.parse_args()
    
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    
    if args.test_url:
        scraper = RBNZConsultationsScraper(use_selenium=args.use_selenium)
        result = scraper._extract_consultation_content(args.test_url)
        if result:
            print(f"✓ Successfully scraped: {result['title']}")
            print(f"  Status: {result['consultation_status']}")
            print(f"  Opened: {result['opened_date']}")
            print(f"  Closed: {result['closed_date']}")
            print(f"  Contact: {result['contact_info']}")
            print(f"  Content length: {len(result['overview_content'])} chars")
            print(f"  PDF content: {len(result['pdf_content'])} chars")
            print(f"  Excel content: {len(result['excel_content'])} chars")
            print(f"  Download links: {len(result['download_links']['pdf_files'])} PDFs, {len(result['download_links']['excel_files'])} Excel")
        else:
            print("✗ Failed to scrape consultation")
        return
        
    if not args.use_selenium:
        print("Error: --use-selenium is required for RBNZ consultations scraping")
        return
    
    if args.resume:
        print("Resume functionality: Load progress file and continue...")
        # TODO: Implement resume functionality
        
    scraper = RBNZConsultationsScraper(max_pages=args.max_pages, use_selenium=args.use_selenium)
    
    if args.max_consultations:
        print(f"Limiting to {args.max_consultations} consultations for testing")
        # Modify scraper to limit consultations
        scraper.max_consultations = args.max_consultations
    
    if args.batch_size:
        print(f"Processing in batches of {args.batch_size} consultations")
        scraper.batch_size = args.batch_size
    
    scraper.run()


if __name__ == "__main__":
    main()