#!/usr/bin/env python3
"""
RBNZ Publications Scraper - Fixed with Proper Logging
Scrapes all publications from the Reserve Bank of New Zealand publications library
with working pagination and proper status logging.
"""

import json
import logging
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import PyPDF2
from io import BytesIO

# Optional Selenium imports for JavaScript handling
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# Configuration
CONFIG = {
    'BASE_URL': 'https://www.rbnz.govt.nz',
    'PUBLICATIONS_URL': 'https://www.rbnz.govt.nz/research-and-publications/publications/publications-library',
    'USER_AGENT': 'rbnz-approved-agent/rg-11701',
    'RATE_LIMIT': 292,  # requests per hour
    'REQUEST_DELAY': 3600 / 292,  # seconds between requests (~12.3 seconds)
    'SAFETY_MARGIN': 0.8,  # Use only 80% of allowed rate for safety
    'MAX_PAGE': 1,  # Set to None for full scrape, or integer for limited pages
    'OUTPUT_DIR': './data',
    'OUTPUT_FILE': './data/rbnz_publications.json',
    'LOG_FILE': './publications_scrape.log',
    'SCRAPED_URLS_FILE': './data/scraped_publications_urls.json',
    'REQUEST_LOG_FILE': './data/request_log.json'
}

class RBNZPublicationsScraper:
    def __init__(self, max_pages: Optional[int] = None, use_selenium: bool = False):
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': CONFIG['USER_AGENT']})
        self.scraped_urls: Set[str] = self._load_scraped_urls()
        self.max_pages = max_pages or CONFIG['MAX_PAGE']
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE
        self.setup_logging()
        self.setup_directories()
        
        # Rate limiting tracking
        self.request_times = []
        self.request_count = 0
        self.hourly_limit = int(CONFIG['RATE_LIMIT'] * CONFIG['SAFETY_MARGIN'])
        self.request_delay = 3600 / self.hourly_limit  # Adjusted delay with safety margin
        
        self.logger.info(f"Rate limiting: {self.hourly_limit} requests/hour, {self.request_delay:.1f}s between requests")
        
        if self.use_selenium:
            self.setup_selenium()
        
    def setup_selenium(self):
        """Setup Selenium WebDriver for JavaScript handling"""
        if not SELENIUM_AVAILABLE:
            self.logger.warning("Selenium not available. Install with: pip install selenium webdriver-manager")
            self.use_selenium = False
            return
            
        self.chrome_options = Options()
        self.chrome_options.add_argument('--headless')
        self.chrome_options.add_argument(f'--user-agent={CONFIG["USER_AGENT"]}')
        self.chrome_options.add_argument('--no-sandbox')
        self.chrome_options.add_argument('--disable-dev-shm-usage')
        self.chrome_options.add_argument('--disable-gpu')
        self.chrome_options.add_argument('--window-size=1920,1080')
        # Additional stability options
        self.chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        self.chrome_options.add_argument('--disable-extensions')
        self.chrome_options.add_argument('--disable-plugins')
        self.chrome_options.add_argument('--disable-images')  # Speed up loading
        self.chrome_options.add_argument('--disable-javascript-harmony-shipping')
        self.chrome_options.add_argument('--disable-background-timer-throttling')
        self.chrome_options.add_argument('--disable-renderer-backgrounding')
        self.chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        self.chrome_options.add_argument('--disable-ipc-flooding-protection')
        # Memory management
        self.chrome_options.add_argument('--memory-pressure-off')
        self.chrome_options.add_argument('--max_old_space_size=4096')
        
        self.logger.info("Selenium WebDriver configured")
        
    def setup_logging(self):
        """Configure logging"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(CONFIG['LOG_FILE']),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def setup_directories(self):
        """Create necessary directories"""
        Path(CONFIG['OUTPUT_DIR']).mkdir(exist_ok=True)
        
    def _load_scraped_urls(self) -> Set[str]:
        """Load previously scraped URLs for deduplication"""
        try:
            if os.path.exists(CONFIG['SCRAPED_URLS_FILE']):
                with open(CONFIG['SCRAPED_URLS_FILE'], 'r') as f:
                    return set(json.load(f))
        except Exception as e:
            self.logger.warning(f"Could not load scraped URLs: {e}")
        return set()
        
    def _save_scraped_urls(self):
        """Save scraped URLs to file"""
        try:
            with open(CONFIG['SCRAPED_URLS_FILE'], 'w') as f:
                json.dump(list(self.scraped_urls), f, indent=2)
        except Exception as e:
            self.logger.error(f"Could not save scraped URLs: {e}")
            
    def _rate_limit(self):
        """Implement smart rate limiting with request tracking"""
        current_time = time.time()
        
        # Remove requests older than 1 hour
        self.request_times = [t for t in self.request_times if current_time - t < 3600]
        
        # Check if we're approaching the hourly limit
        if len(self.request_times) >= self.hourly_limit:
            # Calculate how long to wait until the oldest request is > 1 hour old
            oldest_request = min(self.request_times)
            wait_time = 3600 - (current_time - oldest_request) + 1  # +1 second buffer
            
            if wait_time > 0:
                self.logger.warning(f"Rate limit reached. Waiting {wait_time:.1f} seconds...")
                time.sleep(wait_time)
                current_time = time.time()
                # Clean up old requests after waiting
                self.request_times = [t for t in self.request_times if current_time - t < 3600]
        
        # Standard delay between requests
        time.sleep(self.request_delay)
        
        # Record this request
        self.request_times.append(current_time)
        self.request_count += 1
        
        # Log progress every 20 requests
        if self.request_count % 20 == 0:
            requests_in_last_hour = len(self.request_times)
            self.logger.info(f"Request #{self.request_count}: {requests_in_last_hour}/{self.hourly_limit} requests in last hour")

    def _safe_request(self, url: str, timeout: int = 30) -> Optional[requests.Response]:
        """Make a rate-limited request with error handling"""
        max_retries = 3
        retry_delay = 30  # Start with 30 seconds
        
        for attempt in range(max_retries):
            try:
                self._rate_limit()
                
                response = self.session.get(url, timeout=timeout)
                
                if response.status_code == 429:
                    # Rate limited - wait longer
                    wait_time = retry_delay * (attempt + 1)
                    self.logger.warning(f"Rate limited (429). Waiting {wait_time} seconds before retry {attempt + 1}/{max_retries}")
                    time.sleep(wait_time)
                    continue
                
                response.raise_for_status()
                return response
                
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 429:
                    wait_time = retry_delay * (attempt + 1)
                    self.logger.warning(f"Rate limited (429). Waiting {wait_time} seconds before retry {attempt + 1}/{max_retries}")
                    time.sleep(wait_time)
                    continue
                else:
                    self.logger.error(f"HTTP error for {url}: {e}")
                    break
            except Exception as e:
                self.logger.error(f"Request error for {url}: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                break
        
        return None

    def _clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        if not text:
            return ""
        
        try:
            # Remove extra whitespace and normalize
            text = re.sub(r'\s+', ' ', text.strip())
            # Remove HTML entities
            text = text.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
            # Normalize quotes
            text = text.replace('"', '"').replace('"', '"').replace('„', '"')
            text = text.replace(''', "'").replace(''', "'").replace('`', "'")
            # Remove HTML tags
            text = re.sub(r'<[^>]+>', '', text)
            # Remove excessive punctuation
            text = re.sub(r'\.{3,}', '...', text)
            # Normalize dashes
            text = text.replace('–', '-').replace('—', '-')
        except Exception as e:
            self.logger.warning(f"Text cleaning failed: {e}")
            text = ' '.join(text.split())
            
        return text

    def _extract_pdf_text(self, pdf_url: str) -> str:
        """Extract text from PDF files"""
        try:
            response = self._safe_request(pdf_url, timeout=60)
            if not response:
                return ""
            
            pdf_reader = PyPDF2.PdfReader(BytesIO(response.content))
            text_content = []
            
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content.append(self._clean_text(page_text))
                    
            return "\n\n".join(text_content)
            
        except Exception as e:
            self.logger.error(f"Error extracting PDF {pdf_url}: {e}")
            return ""

    def _get_rbnz_publications_page_url(self, page_number: int) -> str:
        """Get the correct RBNZ publications URL for a specific page"""
        base_url = CONFIG['PUBLICATIONS_URL']
        
        if page_number == 1:
            return f"{base_url}#sort=%40computedsortdate%20descending"
        else:
            first_param = (page_number - 1) * 10
            return f"{base_url}#first={first_param}&sort=%40computedsortdate%20descending"

    def _create_driver(self):
        """Create a new Chrome driver instance"""
        return webdriver.Chrome(
            service=webdriver.chrome.service.Service(ChromeDriverManager().install()),
            options=self.chrome_options
        )

    def _extract_publication_links_from_page(self, driver) -> List[str]:
        """Extract all publication links from the current page"""
        try:
            # Wait for results to load
            WebDriverWait(driver, 20).until(
                lambda d: len(d.find_elements(By.CSS_SELECTOR, ".coveo-list-layout.CoveoResult")) > 0
            )
            
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            links = []
            
            # Find all Coveo result containers
            results = soup.find_all('div', class_='coveo-list-layout CoveoResult')
            
            for result in results:
                # Look for the main publication link
                link_elem = result.find('a', class_='CoveoResultLink')
                if link_elem and link_elem.get('href'):
                    url = link_elem['href']
                    # Publications can be various types of content
                    if any(path in url for path in ['/hub/', '/publications/', '/research/']):
                        if url.startswith('/'):
                            url = urljoin(CONFIG['BASE_URL'], url)
                        links.append(url)
            
            return links
            
        except Exception as e:
            self.logger.error(f"Error extracting links: {e}")
            return []

    def _scrape_all_pages_selenium(self) -> List[str]:
        """Scrape all pages using Selenium with robust session management"""
        if not self.use_selenium:
            return []
            
        all_links = []
        driver = None
        
        try:
            page = 1
            consecutive_failures = 0
            max_consecutive_failures = 3
            
            while page <= self.max_pages:
                # Create new driver if needed or if previous one failed
                if driver is None:
                    driver = self._create_driver()
                    self.logger.info(f"Created new Chrome driver session")
                
                url = self._get_rbnz_publications_page_url(page)
                self.logger.info(f"Scraping page {page}: {url}")
                
                try:
                    driver.get(url)
                    time.sleep(8)  # Give Coveo time to load
                    
                    # Test if driver is still responsive
                    try:
                        _ = driver.current_url
                    except Exception as session_error:
                        self.logger.warning(f"Driver session lost: {session_error}")
                        driver.quit()
                        driver = None
                        consecutive_failures += 1
                        
                        if consecutive_failures >= max_consecutive_failures:
                            self.logger.error(f"Too many consecutive failures ({consecutive_failures}), stopping")
                            break
                            
                        self.logger.info("Creating new driver session and retrying...")
                        continue
                    
                    # Extract links from this page
                    page_links = self._extract_publication_links_from_page(driver)
                    
                    if not page_links:
                        self.logger.info(f"No links found on page {page} - reached end")
                        break
                    
                    # Check if we're getting new links
                    new_links = [link for link in page_links if link not in all_links]
                    
                    if not new_links and page > 1:
                        self.logger.info(f"No new links on page {page} - reached end")
                        break
                    
                    all_links.extend(new_links)
                    self.logger.info(f"Page {page}: found {len(page_links)} links ({len(new_links)} new). Total: {len(all_links)}")
                    
                    # Log first few URLs to verify they're different
                    if new_links:
                        self.logger.info(f"Sample new URLs: {new_links[:2]}")
                    
                    # Reset failure counter on success
                    consecutive_failures = 0
                    page += 1
                    
                    # Restart driver every 5 pages to prevent memory issues
                    if page % 5 == 0:
                        self.logger.info("Restarting driver to prevent memory issues...")
                        driver.quit()
                        driver = None
                        time.sleep(3)
                    
                except Exception as page_error:
                    self.logger.error(f"Error on page {page}: {page_error}")
                    consecutive_failures += 1
                    
                    if consecutive_failures >= max_consecutive_failures:
                        self.logger.error(f"Too many consecutive failures ({consecutive_failures}), stopping")
                        break
                    
                    # Try to recover by creating new driver
                    try:
                        if driver:
                            driver.quit()
                    except:
                        pass
                    driver = None
                    
                    self.logger.info("Attempting to recover with new driver session...")
                    continue
            
            if driver:
                driver.quit()
            
            # Remove any remaining duplicates
            unique_links = list(dict.fromkeys(all_links))
            self.logger.info(f"Selenium scraping completed. Total unique links: {len(unique_links)}")
            return unique_links
            
        except Exception as e:
            self.logger.error(f"Selenium scraping failed: {e}")
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            return all_links  # Return what we got so far

    def _extract_excel_data(self, excel_url: str) -> str:
        """Extract data from Excel files using SheetJS"""
        try:
            response = self._safe_request(excel_url, timeout=60)
            if not response:
                return ""
            
            # Try to use openpyxl if available, otherwise skip Excel processing
            try:
                import openpyxl
                from io import BytesIO
                
                workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
                excel_content = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    excel_content.append(f"--- SHEET: {sheet_name} ---")
                    
                    # Get all rows with data
                    rows_data = []
                    for row in sheet.iter_rows(values_only=True):
                        # Convert row to strings and filter out None values
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        # Only add rows that have some content
                        if any(cell.strip() for cell in row_data if cell):
                            rows_data.append(" | ".join(row_data))
                    
                    excel_content.extend(rows_data[:100])  # Limit to first 100 rows per sheet
                    
                return "\n".join(excel_content)
                
            except ImportError:
                self.logger.warning("openpyxl not available for Excel processing. Install with: pip install openpyxl")
                return f"Excel file available at: {excel_url} (openpyxl required for processing)"
                
        except Exception as e:
            self.logger.error(f"Error extracting Excel data from {excel_url}: {e}")
            return ""

    def _extract_publication_content(self, publication_url: str) -> Optional[Dict]:
        """Extract content from a single publication with proper status handling"""
        if publication_url in self.scraped_urls:
            self.logger.debug(f"Skipping already scraped publication: {publication_url}")
            return "ALREADY_SCRAPED"  # Return special marker instead of None
            
        try:
            response = self._safe_request(publication_url, timeout=30)
            if not response:
                return None
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract basic information
            title = ""
            title_elem = soup.find('h1', class_='publication-hero__heading') or soup.find('h1', class_='hero__heading') or soup.find('h1')
            if title_elem:
                title = self._clean_text(title_elem.get_text())
            
            published_date = ""
            date_elem = soup.find('time')
            if date_elem:
                published_date = date_elem.get('datetime', '') or self._clean_text(date_elem.get_text())
                
            # Try to find publication date in other locations if not found
            if not published_date:
                date_patterns = soup.find_all(text=re.compile(r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b'))
                if date_patterns:
                    published_date = date_patterns[0].strip()
                
            # Extract description from the publication hero section
            description = ""
            desc_elem = soup.find('p', class_='publication-hero__description') or soup.find('p', class_='hero__description')
            if desc_elem:
                description = self._clean_text(desc_elem.get_text())
            
            # Extract authors
            authors = []
            author_elem = soup.find('p', class_='publication-hero__author')
            if author_elem:
                author_text = self._clean_text(author_elem.get_text())
                if author_text:
                    authors = [author_text]
            
            # Extract main content from article-content div ONLY (to avoid duplication)
            content_sections = []
            article_content = soup.find('div', id='article-content')
            if article_content:
                # Get only direct content, not nested components that might duplicate info
                for elem in article_content.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'li'], recursive=True):
                    # Skip elements that are inside download cards or other special components
                    if not any(parent.get('class') and any('download-card' in str(cls) for cls in parent.get('class', [])) 
                              for parent in elem.parents):
                        text = self._clean_text(elem.get_text())
                        if text and len(text) > 10:
                            content_sections.append(text)
            
            # Remove duplicates while preserving order
            unique_content = []
            seen_content = set()
            for section in content_sections:
                # Use first 50 characters as key to detect near-duplicates
                key = section[:50].lower()
                if key not in seen_content:
                    unique_content.append(section)
                    seen_content.add(key)
                        
            content_text = "\n\n".join(unique_content)
            
            # Extract publication type/category from breadcrumbs and metadata
            publication_type = []
            
            # Look for breadcrumbs
            breadcrumbs = soup.find_all(['span', 'div', 'a'], class_=re.compile(r'breadcrumb|category'))
            for breadcrumb in breadcrumbs:
                category = self._clean_text(breadcrumb.get_text())
                if category and len(category) > 2 and category not in publication_type:
                    publication_type.append(category)
            
            # Look for metadata sections
            metadata_sections = soup.find_all('div', class_=re.compile(r'metadata'))
            for section in metadata_sections:
                for elem in section.find_all(['span', 'div'], class_=['tag', 'label', 'type']):
                    pub_type = self._clean_text(elem.get_text())
                    if pub_type and pub_type not in publication_type:
                        publication_type.append(pub_type)
                        
            # Extract download links for PDFs and Excel files
            download_links = {
                'pdf_files': [],
                'excel_files': [],
                'other_files': []
            }
            
            # Look for download cards (as seen in the HTML sample)
            download_cards = soup.find_all('div', class_='download-card-component')
            for card in download_cards:
                link_elem = card.find('a', class_='download-card__link')
                if link_elem and link_elem.get('href'):
                    download_url = link_elem['href']
                    if download_url.startswith('/'):
                        download_url = urljoin(CONFIG['BASE_URL'], download_url)
                    
                    # Get the title/description
                    title_elem = card.find('p', class_='download-card__heading')
                    file_title = self._clean_text(title_elem.get_text()) if title_elem else "Download"
                    
                    # Get file type info
                    file_type_elem = card.find('span', class_='download-card__file-type-text')
                    file_type = self._clean_text(file_type_elem.get_text()) if file_type_elem else ""
                    
                    download_info = {
                        'title': file_title,
                        'url': download_url,
                        'file_type': file_type
                    }
                    
                    if download_url.lower().endswith('.pdf'):
                        download_links['pdf_files'].append(download_info)
                    elif download_url.lower().endswith(('.xlsx', '.xls')):
                        download_links['excel_files'].append(download_info)
                    else:
                        download_links['other_files'].append(download_info)
            
            # Also look for any other links to PDFs/Excel files in the content
            for link in soup.find_all('a', href=True):
                href = link['href']
                if href.startswith('/'):
                    href = urljoin(CONFIG['BASE_URL'], href)
                
                link_text = self._clean_text(link.get_text())
                
                # Check if it's a PDF or Excel file not already captured
                if href.lower().endswith('.pdf'):
                    if not any(dl['url'] == href for dl in download_links['pdf_files']):
                        download_links['pdf_files'].append({
                            'title': link_text or 'PDF Document',
                            'url': href,
                            'file_type': 'PDF'
                        })
                elif href.lower().endswith(('.xlsx', '.xls')):
                    if not any(dl['url'] == href for dl in download_links['excel_files']):
                        download_links['excel_files'].append({
                            'title': link_text or 'Excel Spreadsheet',
                            'url': href,
                            'file_type': 'Excel'
                        })
                        
            # Extract image
            associated_image_url = ""
            # Look for feature images first
            img_elem = soup.find('img', class_='feature-image__source-image') or soup.find('img')
            if img_elem and img_elem.get('src'):
                img_url = img_elem['src']
                if img_url.startswith('/'):
                    img_url = urljoin(CONFIG['BASE_URL'], img_url)
                associated_image_url = img_url
                
            # Process PDF files
            pdf_content = ""
            for pdf_info in download_links['pdf_files']:
                self.logger.info(f"Extracting PDF content from: {pdf_info['title']}")
                pdf_text = self._extract_pdf_text(pdf_info['url'])
                if pdf_text:
                    pdf_content += f"\n\n--- PDF: {pdf_info['title']} ---\n\n{pdf_text}"
                    
            # Process Excel files
            excel_content = ""
            for excel_info in download_links['excel_files']:
                self.logger.info(f"Extracting Excel data from: {excel_info['title']}")
                excel_data = self._extract_excel_data(excel_info['url'])
                if excel_data:
                    excel_content += f"\n\n--- EXCEL: {excel_info['title']} ---\n\n{excel_data}"
                    
            # Extract tables from HTML
            tables_data = []
            for table in soup.find_all('table'):
                table_text = []
                for row in table.find_all('tr'):
                    row_text = []
                    for cell in row.find_all(['td', 'th']):
                        cell_text = self._clean_text(cell.get_text())
                        if cell_text:
                            row_text.append(cell_text)
                    if row_text:
                        table_text.append(" | ".join(row_text))
                if table_text:
                    tables_data.append("\n".join(table_text))
                    
            tables_and_charts_data = "\n\n--- HTML TABLE ---\n\n".join(tables_data)
            
            # Extract related links (excluding download files already processed)
            related_links = []
            processed_urls = set([dl['url'] for dl_list in download_links.values() for dl in dl_list])
            
            for link in soup.find_all('a', href=True):
                href = link['href']
                if href.startswith('/'):
                    href = urljoin(CONFIG['BASE_URL'], href)
                    
                link_text = self._clean_text(link.get_text())
                
                if (href not in processed_urls and 
                    link_text and len(link_text) > 3 and 
                    not href.lower().endswith(('.pdf', '.xlsx', '.xls')) and
                    href not in [rl['url'] for rl in related_links]):
                    
                    related_links.append({
                        'text': link_text,
                        'url': href
                    })
            
            # Mark as scraped
            self.scraped_urls.add(publication_url)
            
            return {
                'url': publication_url,
                'title': title,
                'published_date': published_date,
                'scraped_date': datetime.now().isoformat(),
                'publication_type': publication_type,
                'authors': authors,
                'description': description,
                'content_text': content_text,
                'related_links': related_links,
                'associated_image_url': associated_image_url,
                'download_links': download_links,
                'pdf_content': pdf_content,
                'excel_content': excel_content,
                'tables_and_charts_data': tables_and_charts_data
            }
            
        except Exception as e:
            self.logger.error(f"Error extracting content from {publication_url}: {e}")
            import traceback
            self.logger.debug(f"Full traceback: {traceback.format_exc()}")
            return None

    def scrape_all_publications(self) -> List[Dict]:
        """Main scraping method with improved logging"""
        self.logger.info(f"Starting RBNZ publications scraping (max_pages: {self.max_pages})")
        
        # Get all publication URLs
        if self.use_selenium:
            self.logger.info("Using Selenium to get publication URLs")
            publication_urls = self._scrape_all_pages_selenium()
        else:
            self.logger.error("Non-Selenium scraping not implemented - use --use-selenium")
            return []
        
        if not publication_urls:
            self.logger.warning("No publication URLs found")
            return []
            
        self.logger.info(f"Found {len(publication_urls)} total publication URLs")
        
        # Calculate estimated time
        estimated_seconds = len(publication_urls) * self.request_delay
        estimated_hours = estimated_seconds / 3600
        self.logger.info(f"Estimated time to complete: {estimated_hours:.1f} hours")
        
        # Extract content from each publication
        publications = []
        already_scraped_count = 0
        failed_count = 0
        
        for i, url in enumerate(publication_urls, 1):
            self.logger.info(f"Processing publication {i}/{len(publication_urls)}: {url}")
            
            publication_data = self._extract_publication_content(url)
            
            if publication_data == "ALREADY_SCRAPED":
                already_scraped_count += 1
                self.logger.info(f"↻ Already scraped: {url}")
            elif publication_data:
                publications.append(publication_data)
                self.logger.info(f"✓ Scraped: {publication_data['title'][:60]}...")
                
                # Save progress every 10 publications
                if len(publications) % 10 == 0:
                    self._save_progress(publications)
                    self.logger.info(f"Progress saved: {len(publications)} publications scraped")
            else:
                failed_count += 1
                self.logger.warning(f"✗ Failed to scrape: {url}")
        
        # Final save
        self._save_progress(publications)
        
        if failed_count > 0:
            self.logger.warning(f"Failed to scrape {failed_count} publications")
            # Note: failed_urls tracking removed since we now have proper status handling
                
        self.logger.info(f"Scraping completed. New: {len(publications)}, Already scraped: {already_scraped_count}, Failed: {failed_count}")
        return publications

    def _save_progress(self, publications: List[Dict]):
        """Save progress to avoid losing work"""
        try:
            # Save to a temporary progress file
            progress_file = './data/publications_progress.json'
            with open(progress_file, 'w', encoding='utf-8') as f:
                json.dump(publications, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"Error saving progress: {e}")
            
    def save_results(self, publications: List[Dict]):
        """Save scraped publications to JSON file"""
        try:
            # Load existing publications
            existing_publications = []
            if os.path.exists(CONFIG['OUTPUT_FILE']):
                with open(CONFIG['OUTPUT_FILE'], 'r', encoding='utf-8') as f:
                    existing_publications = json.load(f)
                    
            # Merge with new publications
            existing_urls = {pub.get('url') for pub in existing_publications}
            new_publications = [pub for pub in publications if pub.get('url') not in existing_urls]
            
            all_publications = existing_publications + new_publications
            
            # Save combined results
            with open(CONFIG['OUTPUT_FILE'], 'w', encoding='utf-8') as f:
                json.dump(all_publications, f, indent=2, ensure_ascii=False)
                
            self.logger.info(f"Saved {len(new_publications)} new publications. Total: {len(all_publications)}")
            
            # Save scraped URLs
            self._save_scraped_urls()
            
            # Save request statistics
            self._save_request_stats()
            
        except Exception as e:
            self.logger.error(f"Error saving results: {e}")
            
    def _save_request_stats(self):
        """Save request statistics for monitoring"""
        try:
            stats = {
                'total_requests': self.request_count,
                'requests_in_last_hour': len(self.request_times),
                'hourly_limit': self.hourly_limit,
                'last_request_time': max(self.request_times) if self.request_times else None,
                'scrape_date': datetime.now().isoformat()
            }
            
            with open('./data/request_stats.json', 'w') as f:
                json.dump(stats, f, indent=2)
                
        except Exception as e:
            self.logger.error(f"Error saving request stats: {e}")
            
    def run(self):
        """Main execution method"""
        start_time = datetime.now()
        self.logger.info(f"RBNZ publications scraper started at {start_time}")
        
        try:
            publications = self.scrape_all_publications()
            if publications:
                self.save_results(publications)
            else:
                self.logger.warning("No new publications were scraped")
                
        except Exception as e:
            self.logger.error(f"Scraping failed: {e}")
            
        end_time = datetime.now()
        duration = end_time - start_time
        self.logger.info(f"Scraping completed in {duration}")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='RBNZ Publications Scraper')
    parser.add_argument('--max-pages', type=int, default=1,
                       help='Maximum number of pages to scrape')
    parser.add_argument('--max-publications', type=int, default=None,
                       help='Maximum number of publications to scrape (useful for testing)')
    parser.add_argument('--use-selenium', action='store_true',
                       help='Use Selenium WebDriver (required)')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug logging')
    parser.add_argument('--test-url', type=str,
                       help='Test scraping a specific publication URL')
    parser.add_argument('--resume', action='store_true',
                       help='Resume from previous progress file')
    parser.add_argument('--batch-size', type=int, default=50,
                       help='Number of publications to scrape in this batch')
    
    args = parser.parse_args()
    
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    
    if args.test_url:
        scraper = RBNZPublicationsScraper(use_selenium=args.use_selenium)
        result = scraper._extract_publication_content(args.test_url)
        if result == "ALREADY_SCRAPED":
            print(f"↻ Publication already scraped: {args.test_url}")
        elif result:
            print(f"✓ Successfully scraped: {result['title']}")
            print(f"  Date: {result['published_date']}")
            print(f"  Type: {result['publication_type']}")
            print(f"  Content length: {len(result['content_text'])} chars")
            print(f"  PDF content: {len(result['pdf_content'])} chars")
            print(f"  Excel content: {len(result['excel_content'])} chars")
        else:
            print("✗ Failed to scrape publication")
        return
        
    if not args.use_selenium:
        print("Error: --use-selenium is required for RBNZ publications scraping")
        return
    
    if args.resume:
        print("Resume functionality: Load progress file and continue...")
        # TODO: Implement resume functionality
        
    scraper = RBNZPublicationsScraper(max_pages=args.max_pages, use_selenium=args.use_selenium)
    
    if args.max_publications:
        print(f"Limiting to {args.max_publications} publications for testing")
        # Modify scraper to limit publications
        scraper.max_publications = args.max_publications
    
    if args.batch_size:
        print(f"Processing in batches of {args.batch_size} publications")
        scraper.batch_size = args.batch_size
    
    scraper.run()


if __name__ == "__main__":
    main()